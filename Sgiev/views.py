from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth import authenticate, logout
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.core.paginator import Paginator
from django.db import models, transaction
from django.http import JsonResponse
from datetime import datetime, date, timedelta
from decimal import Decimal, InvalidOperation
from .models import (
    Categoria, 
    Proveedor,
    Usuarios, 
    Producto, 
    Venta, 
    Venta_has_producto, 
    Movimiento_inventario,
    Envio, 
    Mensajeria,
    Compra_proveedor,
    Compra_detalle
)
from .forms import LoginForm, UsuarioForm, VentaForm, AgregarProductoForm,EditarEstadoVentaForm,EnvioForm, MensajeriaForm,EnvioEditarOperarioForm
from .decorators import admin_required
import json
from django.utils.safestring import mark_safe
from django.template.loader import render_to_string
import openpyxl
from openpyxl.utils import get_column_letter
from django.http import HttpResponse
from django.db.models import Q
from django.db.models import F 
from django.template.loader import get_template
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import letter
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph
from reportlab.lib import colors
from django.views.decorators.cache import never_cache
from django.http import HttpResponseRedirect 
from django.urls import reverse
from django.utils import timezone
from django.db.models import Sum, Count, Q, F,Min
from django.utils import timezone
from datetime import datetime, date, timedelta
from decimal import Decimal
import json
from django.db import IntegrityError, transaction
import uuid
from uuid import uuid4
from django.db.models import Sum, F, Q, Count, Avg
from django.db.models.functions import Coalesce





def index(request):  
    """
    Vista principal - Landing page
    """
    if request.user.is_authenticated:
        return redirect('dashboard')
    
    return render(request, 'index.html')


def login(request):
    return render(request, 'login.html')


def admin(request):
    return render(request, 'admin.html')

#dashboard
@login_required(login_url='login')
def dashboard_view(request):
    usuario = request.user
    es_admin = usuario.tipo_usu == 'administrador'

    # Fecha actual y rango del mes
    hoy = timezone.now()
    inicio_mes = hoy.replace(day=1, hour=0, minute=0, second=0, microsecond=0)

    # ===== FILTRAR DATOS SEGÚN EL ROL =====
    if es_admin:
        ventas_query = Venta.objects.all()
        envios_query = Envio.objects.all()
    else:
        ventas_query = Venta.objects.filter(usuarios_id_usuario=usuario)
        envios_query = Envio.objects.filter(usuarios_id_usuario=usuario)

    # ===== ESTADÍSTICAS DE VENTAS =====
    ventas_mes = ventas_query.filter(fecha_factura__gte=inicio_mes)

    # CALCULAR INGRESOS REALES (solo lo abonado)
    ingresos_reales_mes = Decimal('0')
    for venta in ventas_mes:
        if venta.estado_pago == 'pagado':
            ingresos_reales_mes += venta.valor_total
        elif venta.estado_pago == 'parcial':
            ingresos_reales_mes += venta.abono  # Solo lo abonado
        # 'pendiente' no suma nada

    # VENTAS TOTALES (históricas, solo lo realmente cobrado)
    ventas_totales = Decimal('0')
    for venta in ventas_query.all():
        if venta.estado_pago == 'pagado':
            ventas_totales += venta.valor_total
        elif venta.estado_pago == 'parcial':
            ventas_totales += venta.abono

    # PRODUCTOS VENDIDOS DEL MES
    try:
        productos_vendidos = Venta_has_producto.objects.filter(
            venta_idfactura__in=ventas_mes,
            venta_idfactura__fecha_factura__gte=inicio_mes
        ).aggregate(total=Sum('cantidad'))['total'] or 0
    except Exception:
        productos_vendidos = 0

    # Mes anterior para comparación
    inicio_mes_anterior = (inicio_mes - timedelta(days=1)).replace(day=1)
    ventas_mes_anterior_query = ventas_query.filter(
        fecha_factura__gte=inicio_mes_anterior,
        fecha_factura__lt=inicio_mes
    )

    ingresos_mes_anterior = Decimal('0')
    for venta in ventas_mes_anterior_query:
        if venta.estado_pago == 'pagado':
            ingresos_mes_anterior += venta.valor_total
        elif venta.estado_pago == 'parcial':
            ingresos_mes_anterior += venta.abono

    if ingresos_mes_anterior > 0:
        ingresos_porcentaje = round(
            ((ingresos_reales_mes - ingresos_mes_anterior) / ingresos_mes_anterior) * 100, 1
        )
    else:
        ingresos_porcentaje = 100 if ingresos_reales_mes > 0 else 0

    # ===== DATOS DEL GRÁFICO (últimos 6 meses con ingresos reales) =====
    meses_labels = []
    meses_valores = []

    meses = ['Ene', 'Feb', 'Mar', 'Abr', 'May', 'Jun', 'Jul', 'Ago', 'Sep', 'Oct', 'Nov', 'Dic']

    for i in range(5, -1, -1):
        fecha_mes = hoy - timedelta(days=30 * i)
        inicio = fecha_mes.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
        
        if i == 0:
            fin = hoy
        else:
            siguiente_mes = inicio + timedelta(days=32)
            fin = siguiente_mes.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
        
        ventas_rango = ventas_query.filter(
            fecha_factura__gte=inicio,
            fecha_factura__lt=fin
        )
        
        total_mes = Decimal('0')
        for venta in ventas_rango:
            if venta.estado_pago == 'pagado':
                total_mes += venta.valor_total
            elif venta.estado_pago == 'parcial':
                total_mes += venta.abono
        
        mes_nombre = meses[fecha_mes.month - 1]
        meses_labels.append(mes_nombre)
        meses_valores.append(float(total_mes))

    # ===== ESTADÍSTICAS DE ENVÍOS =====
    envios_pendientes = envios_query.filter(estado_envio='pendiente').count()
    envios_transito = envios_query.filter(estado_envio='en_transito').count()

    envios_entregados_mes = envios_query.filter(
        estado_envio='entregado',
        fecha_entrega__gte=inicio_mes
    ).count()

    envios_devueltos_mes = envios_query.filter(
        estado_envio='devuelto',
        fecha_entrega__gte=inicio_mes
    ).count()

    envios_entregados_recientes = envios_query.filter(
        estado_envio='entregado'
    ).select_related('venta_idfactura', 'fk_mensajeria').order_by('-fecha_entrega')[:5]

    # ===== ALERTAS DE INVENTARIO =====
    if es_admin:
        productos_agrupados = (
            Producto.objects.filter(activo=1)
            .values(
                'nombre_producto',
                'descripcion_producto',
                'categoria_idcategoria__nombre_categoria',
            )
            .annotate(
                stock_total=Sum('stock_actual'),
                stock_minimo_ref=Min('stock_minimo'),
            )
            .order_by('nombre_producto')
        )

        productos_stock_bajo = []

        for p in productos_agrupados:
            stock_total = p['stock_total'] or 0
            stock_minimo = p['stock_minimo_ref'] or 0

            if stock_total <= stock_minimo:
                productos_stock_bajo.append({
                    'nombre': p['nombre_producto'],
                    'descripcion': p['descripcion_producto'],
                    'categoria': p['categoria_idcategoria__nombre_categoria'],
                    'stock_total': stock_total,
                    'stock_minimo': stock_minimo,
                })

        productos_stock_bajo = productos_stock_bajo[:10]
    else:
        productos_stock_bajo = []

    # ===== VENTAS RECIENTES =====
    ventas_recientes = list(ventas_query.order_by('-fecha_factura')[:5])
    historial_ventas = list(ventas_query.order_by('-fecha_factura')[:10])

    # ===== ENVÍOS PRÓXIMOS =====
    envios_proximos = list(envios_query.filter(
        estado_envio='en_transito',
        fecha_entrega__lte=hoy + timedelta(days=3)
    ).order_by('fecha_entrega')[:5])

    # ===== CONTEXTO =====
    context = {
        'usuario': usuario,
        'es_admin': es_admin,

        'estadisticas_dashboard': {
            'ingresos_mes': ingresos_reales_mes,
            'ingresos_porcentaje': abs(ingresos_porcentaje),
            'ingresos_crecimiento': ingresos_porcentaje >= 0,
            'ventas_totales': ventas_totales,
            'productos_vendidos': productos_vendidos,
            'envios_pendientes': envios_pendientes,
            'envios_transito': envios_transito,
            'envios_entregados': envios_entregados_mes,
            'envios_devueltos': envios_devueltos_mes,
        },

        'ventas_recientes': ventas_recientes,
        'historial_ventas': historial_ventas,

        'grafico_labels': json.dumps(meses_labels),
        'grafico_valores': json.dumps(meses_valores),

        'productos_stock_bajo': productos_stock_bajo,
        'envios_proximos': envios_proximos,
        'envios_entregados_recientes': list(envios_entregados_recientes),
        'tiene_alertas': len(productos_stock_bajo) > 0 or envios_pendientes > 0,
    }

    return render(request, 'dashboard.html', context)

# CATEGORIA

def inicio_cat(request):
    return render(request, 'categoria/index.html')




def list_categoria(request):
    search = request.GET.get('search', '')

    categorias = Categoria.objects.all()

    # Filtro por búsqueda
    if search:
        categorias = categorias.filter(
            nombre_categoria__icontains=search
        )

    # Paginación (10 por página)
    paginator = Paginator(categorias, 10)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    usuario = request.user                     # <-- usuario logueado
    es_admin = request.user.tipo_usu == 'administrador'  # <-- bandera admin

    context = {
        'page_obj': page_obj,
        'search': search,
        'usuario': usuario,     # <-- añadido para badge
        'es_admin': es_admin    # <-- añadido para badge
    }

    return render(request, 'categoria/index.html', context)




def registro_categoria(request):
    if request.method == "POST":
        nombre = request.POST.get('nombreCat')
        descripcion = request.POST.get('descCat')
        fecha = datetime.now()
        estado = 1

        categoria = Categoria(
            nombre_categoria=nombre,
            descripcion_categoria=descripcion,
            fecha_creacion=fecha,
            activo=estado
        )
        
        categoria.save()
        return redirect('list_categoria')


    usuario = request.user   
    es_admin = request.user.tipo_usu == 'administrador'  

    return render(request, 'categoria/nuevocat.html', {
        'usuario': usuario,     
        'es_admin': es_admin    
    })


def pre_editar_categoria(request, id):
    categoria = Categoria.objects.get(id=id)


    usuario = request.user
    es_admin = request.user.tipo_usu == 'administrador'

    data = {
        'categoria': categoria,
        'usuario': usuario,   
        'es_admin': es_admin  
    }

    return render(request, 'categoria/editarcat.html', data)



def editar_categoria(request, id):
    categoria = Categoria.objects.get(id=id)

    if request.method == "POST":
        categoria.nombre_categoria = request.POST.get('nombreCat')
        categoria.descripcion_categoria = request.POST.get('descCat')
        categoria.activo = request.POST.get('estadoCat')
        categoria.save()

        return redirect("list_categoria")

    return redirect("list_categoria")



def eliminar_categoria(request, id):
    categoria = Categoria.objects.get(id=id)
    categoria.delete()
    return redirect('list_categoria')

#PRODUCTOS


# ===== VISTA ACTUALIZADA: list_producto =====
@login_required
def list_producto(request):
    # Productos agrupados (maestros) con precio de venta del catálogo
    productos_agrupados = Producto.objects.filter(
        activo=1,
        codigo_barras='SIN_LOTE_CATALOGO'  # Solo productos maestros
    ).select_related('categoria_idcategoria').values(
        'id',
        'nombre_producto', 
        'descripcion_producto',
        'categoria_idcategoria__nombre_categoria',
        'precio_venta',  # Precio de venta del maestro
        'precio_compra',
        'margen_ganancia'
    ).annotate(
        stock_total=Coalesce(
            Sum('stock_actual', filter=Q(activo=1)),
            0
        )
    ).order_by('nombre_producto')

    # Calcular stock total real sumando todos los lotes
    productos_final = []
    for producto_maestro in productos_agrupados:
        # Sumar stock de todos los lotes de este producto
        stock_real = Producto.objects.filter(
            nombre_producto=producto_maestro['nombre_producto'],
            descripcion_producto=producto_maestro['descripcion_producto'],
            activo=1
        ).exclude(codigo_barras='SIN_LOTE_CATALOGO').aggregate(
            total=Sum('stock_actual')
        )['total'] or 0
        
        producto_maestro['stock_total'] = stock_real
        productos_final.append(producto_maestro)

    # Lotes para el modal de salida
    producto_lotes_original = Producto.objects.select_related(
        'categoria_idcategoria', 
        'proveedor_idproveedor'
    ).exclude(codigo_barras='SIN_LOTE_CATALOGO').filter(activo=1)
    
    lotes_disponibles = []
    for p in producto_lotes_original:
        nombre = p.nombre_producto.strip()
        descripcion = p.descripcion_producto.strip()
        nombre_completo = f"{nombre} {descripcion}"
        
        lotes_disponibles.append({
            'id': p.id,
            'nombre_producto_completo': nombre_completo,
            'codigo_barras': p.codigo_barras,
            'stock_actual': p.stock_actual,
            'fecha_vencimiento': p.fecha_vencimiento.strftime('%Y-%m-%d') if p.fecha_vencimiento else 'N/A'
        })

    lotes_json_string = json.dumps(lotes_disponibles)
    lotes_json_safe = mark_safe(lotes_json_string)
    
    historial_compras = Compra_proveedor.objects.all().order_by('-fecha_compra')[:10]
    historial_salidas = Movimiento_inventario.objects.filter(
        tipo_movimiento__in=['ajuste', 'venta']
    ).order_by('-fecha_movimiento')[:15]

    usuario = request.user
    es_admin = request.user.tipo_usu == 'administrador'

    data = {
        'producto': productos_final,
        'lotes_json': lotes_json_safe,
        'historial_compras': historial_compras,
        'historial_salidas': historial_salidas,
        'usuario': usuario,
        'es_admin': es_admin,
    }
    
    return render(request, 'producto/index.html', data)

# ===== NUEVA VISTA: Editar Producto Maestro (con precio de venta) =====
@login_required
def editar_producto_maestro(request, id):
    """
    Edita el producto maestro (catálogo), incluyendo precio de venta.
    Solo afecta al registro SIN_LOTE_CATALOGO.
    """
    producto_maestro = get_object_or_404(
        Producto, 
        id=id, 
        codigo_barras='SIN_LOTE_CATALOGO'
    )
    categorias = Categoria.objects.filter(activo=1)
    proveedores = Proveedor.objects.filter(activo=1)

    usuario = request.user
    es_admin = request.user.tipo_usu == 'administrador'

    if request.method == "POST":
        nombre = request.POST.get('nombre_producto', '').strip()
        descripcion = request.POST.get('descripcion_producto', '').strip()
        categoria_id = request.POST.get('categoria_idcategoria')
        proveedor_id = request.POST.get('proveedor_idproveedor')
        
        # Validar unicidad (excluyendo el actual)
        existe_duplicado = Producto.objects.filter(
            nombre_producto=nombre,
            descripcion_producto=descripcion,
            codigo_barras="SIN_LOTE_CATALOGO"
        ).exclude(id=producto_maestro.id).exists()
        
        if existe_duplicado:
            messages.error(
                request, 
                f"Ya existe otro producto maestro con el nombre '{nombre}' y la misma descripción."
            )
            return redirect('editar_producto_maestro', id=id)
        
        # Actualizar datos
        producto_maestro.nombre_producto = nombre
        producto_maestro.descripcion_producto = descripcion
        producto_maestro.registrosanitario = request.POST.get('registrosanitario', '')
        
        try:
            nuevo_precio_compra = Decimal(request.POST.get('precio_compra') or '0')
            nuevo_precio_venta = Decimal(request.POST.get('precio_venta') or '0')
            
            producto_maestro.precio_compra = nuevo_precio_compra
            producto_maestro.precio_venta = nuevo_precio_venta
            
            # Recalcular margen
            if nuevo_precio_compra > 0 and nuevo_precio_venta > 0:
                margen = ((nuevo_precio_venta - nuevo_precio_compra) / nuevo_precio_compra) * Decimal('100.00')
                producto_maestro.margen_ganancia = margen.quantize(Decimal('0.01'))
            else:
                producto_maestro.margen_ganancia = Decimal('0.00')
                
        except (InvalidOperation, ValueError):
            messages.error(request, "Error en formato de precios")
            return redirect('editar_producto_maestro', id=id)
        
        try:
            producto_maestro.stock_minimo = int(request.POST.get('stock_minimo') or 0)
            producto_maestro.stock_maximo = int(request.POST.get('stock_maximo') or 0)
        except:
            pass
        
        producto_maestro.categoria_idcategoria = get_object_or_404(Categoria, id=categoria_id)
        producto_maestro.proveedor_idproveedor = get_object_or_404(Proveedor, id=proveedor_id)
        
        producto_maestro.save()
        
        # Actualizar todos los lotes con el nuevo precio de venta y margen
        Producto.objects.filter(
            nombre_producto=nombre,
            descripcion_producto=descripcion,
            activo=1
        ).exclude(codigo_barras='SIN_LOTE_CATALOGO').update(
            precio_venta=producto_maestro.precio_venta,
            margen_ganancia=producto_maestro.margen_ganancia
        )
        
        messages.success(request, f"Producto maestro '{nombre}' actualizado. Precio de venta aplicado a todos los lotes.")
        return redirect('list_producto')

    data = {
        'producto': producto_maestro,
        'categorias': categorias,
        'proveedores': proveedores,
        'usuario': usuario,
        'es_admin': es_admin,
    }
    return render(request, 'producto/editar_maestro.html', data)

# ===== VISTA ACTUALIZADA: detalle_producto_modal =====
def detalle_producto_modal(request, producto_id):
    # Puede ser un lote o el maestro, obtener el maestro
    lote_o_maestro = get_object_or_404(Producto, pk=producto_id)
    
    nombre_prod_general = lote_o_maestro.nombre_producto
    descripcion_prod_general = lote_o_maestro.descripcion_producto
    
    # Obtener el producto maestro
    try:
        producto_maestro = Producto.objects.get(
            nombre_producto=nombre_prod_general,
            descripcion_producto=descripcion_prod_general,
            codigo_barras='SIN_LOTE_CATALOGO',
            activo=1
        )
    except Producto.DoesNotExist:
        producto_maestro = lote_o_maestro
    
    # Todos los lotes (excluyendo el maestro)
    todos_los_lotes_del_producto = Producto.objects.select_related(
        'categoria_idcategoria', 'proveedor_idproveedor'
    ).filter(
        nombre_producto=nombre_prod_general,
        descripcion_producto=descripcion_prod_general,
        activo=1
    ).exclude(codigo_barras='SIN_LOTE_CATALOGO').order_by('fecha_vencimiento', 'codigo_barras')
    
    stock_total_data = todos_los_lotes_del_producto.aggregate(total_stock=Sum('stock_actual'))
    stock_total = stock_total_data['total_stock'] or 0

    ids_lotes = [l.id for l in todos_los_lotes_del_producto]
    movimientos_producto = Movimiento_inventario.objects.filter(
        producto_idproducto__in=ids_lotes
    ).order_by('-fecha_movimiento')[:10]
    
    context = {
        'producto_maestro': producto_maestro,
        'lotes': todos_los_lotes_del_producto,
        'stock_total': stock_total,
        'movimientos': movimientos_producto,
    }

    return render(request, 'producto/detalle_producto_modal_content.html', context)
@login_required
def registro_producto(request):
    usuario = request.user  
    es_admin = request.user.tipo_usu == 'administrador' 
   
    def render_form_with_data(request, form_data=None):
        categorias = Categoria.objects.filter(activo=1)
        proveedores = Proveedor.objects.filter(activo=1)

        data = {
            'categorias': categorias,
            'proveedores': proveedores,
            'form_data': form_data or {},
            'usuario': usuario, 
            'es_admin': es_admin 
        }
        return render(request, 'producto/nuevoprod.html', data)

    if request.method == "POST":
    
        nombre = request.POST.get('nombre_producto', '').strip()
        descripcion = request.POST.get('descripcion_producto', '').strip()
        categoria_id = request.POST.get('categoria_idcategoria')
        proveedor_id = request.POST.get('proveedor_idproveedor')
        
        registro_sanitario_valor = request.POST.get('registrosanitario', '').strip()
        if not registro_sanitario_valor:
            registro_sanitario_valor = 'SIN_REGISTRO'

        codigo = "SIN_LOTE_CATALOGO"
        activo = 1
        precio_compra = Decimal('0.00')
        precio_venta = Decimal('0.00')
        margen = Decimal('0.00')
        stock_actual = 0
        stock_min = 0
        stock_max = 0
        fecha_ven = date(2099, 12, 31)
        fecha_cre = timezone.now()
        existe_catalogo = Producto.objects.filter(
            nombre_producto=nombre,
            descripcion_producto=descripcion,
            codigo_barras=codigo
        ).exists()
        
        if existe_catalogo:
            messages.error(request, f"Error de Unicidad: Ya existe un producto de catálogo con el nombre '{nombre}' y la misma descripción. (Código: {codigo})")
            return render_form_with_data(request, request.POST)

        try:
            categoria = get_object_or_404(Categoria, id=categoria_id)
            proveedor = get_object_or_404(Proveedor, id=proveedor_id)

            producto = Producto(
                nombre_producto = nombre,
                descripcion_producto = descripcion,
                codigo_barras = codigo,
                precio_compra = precio_compra,
                precio_venta = precio_venta,
                margen_ganancia = margen,
                stock_actual = stock_actual,
                stock_minimo = stock_min,
                stock_maximo = stock_max,
                fecha_vencimiento = fecha_ven,
                fecha_creacion = fecha_cre,
                activo = activo,
                registrosanitario = registro_sanitario_valor,
                categoria_idcategoria = categoria,
                proveedor_idproveedor = proveedor,
            )
            
            producto.save()
            messages.success(request, f"Producto de catálogo '{nombre}' registrado exitosamente.")
            return redirect('list_producto')

        except Exception as e:
            messages.error(request, f"Ocurrió un error inesperado al registrar el producto: {e}")
            return render_form_with_data(request, request.POST)

    return render_form_with_data(request)

# ===== VISTA ACTUALIZADA: Editar Lote Individual =====
@login_required
def pre_editar_producto(request, id):
    """
    Edita un lote individual. Solo permite editar:
    - Stock actual
    - Fecha de vencimiento
    - Stock mínimo/máximo
    NO permite cambiar nombre, descripción, categoría (son del maestro)
    """
    lote = get_object_or_404(Producto, id=id)
    
    # Verificar que NO sea el maestro
    if lote.codigo_barras == 'SIN_LOTE_CATALOGO':
        messages.warning(request, "Para editar el producto maestro, use la opción 'Editar Precio' desde el listado.")
        return redirect('list_producto')
    
    usuario = request.user
    es_admin = request.user.tipo_usu == 'administrador'

    data = {
        'lote': lote,
        'usuario': usuario,
        'es_admin': es_admin
    }
    return render(request, 'producto/editar_lote.html', data)


@login_required
def editar_producto(request, id):
    """
    Guarda cambios de un lote individual
    """
    if request.method == "POST":
        lote = get_object_or_404(Producto, id=id)
        
        # Solo permitir edición de campos específicos del lote
        try:
            lote.stock_actual = int(request.POST.get('stock_actual') or 0)
        except:
            lote.stock_actual = 0
            
        try:
            lote.stock_minimo = int(request.POST.get('stock_minimo') or 0)
        except:
            lote.stock_minimo = 0
            
        try:
            lote.stock_maximo = int(request.POST.get('stock_maximo') or 0)
        except:
            lote.stock_maximo = 0

        fecha_venc = request.POST.get('fecha_vencimiento')
        if fecha_venc:
            lote.fecha_vencimiento = fecha_venc
        
        lote.registrosanitario = request.POST.get('registrosanitario', '')
        
        lote.save()
        messages.success(request, f"Lote '{lote.codigo_barras}' actualizado exitosamente.")

    return redirect('list_producto')

@transaction.atomic
def eliminar_producto(request, id):
    
    usuario_id = request.session.get('_auth_user_id') 
    usuario_actual = None
    
    if usuario_id:
        try:
            usuario_actual = Usuarios.objects.get(id=usuario_id)
        except Usuarios.DoesNotExist:
            pass 
            
 
    producto = get_object_or_404(Producto.objects.select_for_update(), id=id)

    cantidad_eliminada = producto.stock_actual
    
    Movimiento_inventario.objects.create(
        producto_idproducto=producto,
        cantidad=cantidad_eliminada,
        tipo_movimiento='ajuste',   
        stock_anterior=producto.stock_actual,
        stock_nuevo=0, 
        
        precio_unitario=producto.precio_compra,
        valor_total=producto.precio_compra * cantidad_eliminada,
        
        referencia_id=producto.id, 
        tipo_referencia='ajuste', 
        
        observaciones=f'AJUSTE NEGATIVO: Eliminación total. Usuario ID: {usuario_id if usuario_id else "No identificado"}',
        
        usuarios_id_usuario=usuario_actual, 
        fecha_movimiento=timezone.now()
    )
    
    producto.delete()
    
    return redirect('list_producto')

def generar_reporte_productos(request):
    
    categoria_id = request.GET.get('categoria')
    stock_estado = request.GET.get('stock_estado')
    formato = request.GET.get('formato')

    productos_query = Producto.objects.all().order_by('nombre_producto')
    
    productos_query = productos_query.select_related('categoria_idcategoria')

   
    if categoria_id:
    
        if categoria_id.isdigit():
            productos_query = productos_query.filter(categoria_idcategoria__id=categoria_id)

    if stock_estado == 'bajo':
       
        productos_query = productos_query.filter(stock_actual__lte=F('stock_minimo'))
        
    elif stock_estado == 'vencido':
       
        fecha_limite = date.today() + timedelta(days=30)
        productos_query = productos_query.filter(fecha_vencimiento__lte=fecha_limite).order_by('fecha_vencimiento')


 
    if formato == 'excel':
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        response['Content-Disposition'] = 'attachment; filename="reporte_inventario_filtrado.xlsx"'

        workbook = openpyxl.Workbook()
        worksheet = workbook.active
        worksheet.title = "Inventario Romar Natural"

        
        columns = [
            'ID', 'Nombre', 'Descripción', 'Lote', 'Categoría', 
            'Stock Actual', 'Stock Mínimo', 'Precio Venta', 'Vencimiento'
        ]
        
       
        row_num = 1
        for col_num, column_title in enumerate(columns, 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.value = column_title
            
     
        for producto in productos_query:
            row_num += 1
            row = [
                producto.id,
                producto.nombre_producto,
                producto.descripcion_producto,
                producto.codigo_barras,
                producto.categoria_idcategoria.nombre_categoria if producto.categoria_idcategoria else 'Sin Categoría',
                producto.stock_actual,
                producto.stock_minimo,
                producto.precio_venta,
                producto.fecha_vencimiento.strftime("%Y-%m-%d") if producto.fecha_vencimiento else 'N/A'
            ]
            
            for col_num, cell_value in enumerate(row, 1):
                worksheet.cell(row=row_num, column=col_num, value=cell_value)
        
        workbook.save(response)
        return response


    elif formato == 'pdf':
        
       
        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition'] = 'attachment; filename="reporte_inventario_filtrado.pdf"'

       
        doc = SimpleDocTemplate(response, pagesize=letter)
        styles = getSampleStyleSheet()
        story = [] 
        story.append(Paragraph("<b>REPORTE DE INVENTARIO - ROMAR NATURAL</b>", styles['h1']))
        story.append(Paragraph(f"Fecha del Reporte: {date.today().strftime('%Y-%m-%d')}", styles['Normal']))
        
        filtro_desc = "Todos los productos"
        if stock_estado == 'bajo':
            filtro_desc = "Productos con Stock Bajo/Mínimo"
        elif stock_estado == 'vencido':
            filtro_desc = "Productos Próximos a Vencer (30 días)"
            
        story.append(Paragraph(f"Filtro Aplicado: {filtro_desc}", styles['Normal']))
        story.append(Paragraph("<br/>", styles['Normal'])) 
        
      
        data = [
            ['ID', 'Nombre Producto', 'Lote', 'Stock', 'Mínimo', 'Vencimiento', 'Categoría']
        ]
        
        for producto in productos_query:
            data.append([
                producto.id,
                f"{producto.nombre_producto} {producto.descripcion_producto}",
                producto.codigo_barras,
                producto.stock_actual,
                producto.stock_minimo,
                producto.fecha_vencimiento.strftime("%Y-%m-%d") if producto.fecha_vencimiento else 'N/A',
                producto.categoria_idcategoria.nombre_categoria if producto.categoria_idcategoria else 'S/C'
            ])

        
        if not data or len(data) == 1:
            story.append(Paragraph("No se encontraron productos con los filtros seleccionados.", styles['Normal']))
        else:
            table = Table(data, colWidths=[40, 150, 80, 50, 50, 90, 100])
            
            
            table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.blue),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('GRID', (0, 0), (-1, -1), 1, colors.black),
              
                ('BACKGROUND', (0, 0), (-1, -1), colors.white),
            ]))
            story.append(table)

        
        doc.build(story)
        
        return response
        

    return HttpResponse("Formato de reporte no válido.", status=400)


def registrar_salida_inventario_ajuste(request):
    if request.method == "POST":
        
        producto_id_lote = request.POST.get('producto_id_lote') 
        codigo_barras_lote = request.POST.get('codigo_barras_hidden') 
        cantidad_salida = request.POST.get('cantidad_salida')
        motivo_salida = request.POST.get('motivo_salida')

        
        if not producto_id_lote or not producto_id_lote.isdigit():
            messages.error(request, "Error: Producto no seleccionado o ID no válido.")
            return redirect('list_producto')

        try:
            cantidad = int(cantidad_salida)
        except ValueError:
            messages.error(request, "Error: Cantidad inválida.")
            return redirect('list_producto')

        
    
        response_redirect = HttpResponseRedirect(reverse('list_producto'))
        response_redirect['Cache-Control'] = 'no-cache, no-store, must-revalidate' 
    
        try:
            with transaction.atomic():
                
               
                producto = Producto.objects.select_for_update().get(id=producto_id_lote)
                
               
                usuario = None
                if request.user.is_authenticated:
                    try:
                        usuario = Usuarios.objects.get(id=request.user.id) 
                    except Usuarios.DoesNotExist:
                        
                        messages.error(request, "Error de autenticación: El usuario logueado no está registrado en el modelo de Usuarios.")
                        raise Exception("Usuario no encontrado en el modelo Usuarios.") 
                else:
                    
                    messages.error(request, "Error de autenticación: Debe iniciar sesión para registrar movimientos.")
                    raise Exception("Usuario no autenticado.")


                
                if cantidad <= 0 or cantidad > producto.stock_actual:
                    messages.error(request, f"Error: Stock insuficiente. Stock actual del lote {producto.codigo_barras}: {producto.stock_actual}")
                    raise Exception("Stock Insuficiente.")
                
                
                stock_anterior = producto.stock_actual
                precio_unitario = producto.precio_compra 
                valor_total = precio_unitario * cantidad

                Producto.objects.filter(id=producto_id_lote).update(stock_actual=F('stock_actual') - cantidad)

                producto.refresh_from_db()

                stock_final_real = producto.stock_actual
                print(f"DEBUG POST-UPDATE: Stock verificado en DB: {stock_final_real}")
                

                movimiento = Movimiento_inventario(
                    producto_idproducto=producto, 
                    usuarios_id_usuario=usuario, 
                    
                    tipo_movimiento='ajuste', 
                    cantidad=cantidad,
                    stock_anterior=stock_anterior,
                    stock_nuevo=stock_final_real, 
                    precio_unitario=precio_unitario,
                    valor_total=valor_total,
                    
                    referencia_id=0, 
                    tipo_referencia='ajuste', 
                    imagen_comprobante='',
                    
                    observaciones=f"AJUSTE POR BAJA/DEVOLUCIÓN. Motivo: {motivo_salida}. Lote Retirado (Código de Barras): {codigo_barras_lote}", 
                )
                movimiento.save()
                
            messages.success(request, f"Salida de {cantidad} unidades registrada. Nuevo stock: {stock_final_real}")
            return response_redirect 
            
        except Exception as e:
            
            
            if str(e) not in ["Stock Insuficiente.", "Usuario no encontrado en el modelo Usuarios.", "Usuario no autenticado."]:
                 print(f"ERROR FATAL (Rollback): Falló al registrar el ajuste de inventario. CAUSA: {e}")
                 messages.error(request, "Error fatal al procesar el ajuste. El stock no fue modificado.")
            
            return response_redirect 

    return redirect('list_producto')

    
def detalle_producto_modal(request, producto_id):
    
    lote_original = get_object_or_404(Producto, pk=producto_id)
    
    
    nombre_prod_general = lote_original.nombre_producto
    descripcion_prod_general = lote_original.descripcion_producto
    
    
    todos_los_lotes_del_producto = Producto.objects.select_related(
        'categoria_idcategoria', 'proveedor_idproveedor'
    ).filter(
        nombre_producto=nombre_prod_general,
        descripcion_producto=descripcion_prod_general,
        activo=1 
    ).order_by('fecha_vencimiento', 'codigo_barras') 
    
   
    stock_total_data = todos_los_lotes_del_producto.aggregate(total_stock=Sum('stock_actual'))
    stock_total = stock_total_data['total_stock'] or 0


    ids_lotes = [l.id for l in todos_los_lotes_del_producto]
    movimientos_producto = Movimiento_inventario.objects.filter(
        producto_idproducto__in=ids_lotes
    ).order_by('-fecha_movimiento')[:10]
    
    context = {
        'producto_general': lote_original, 
        'lotes': todos_los_lotes_del_producto,
        'stock_total': stock_total,         
        'movimientos': movimientos_producto,
    }

    return render(request, 'producto/detalle_producto_modal_content.html', context)




#VENTA DE PROVEEDOR
def obtener_productos_por_proveedor(request, proveedor_id):
    productos_filtrados = Producto.objects.filter(
        proveedor_idproveedor__id=proveedor_id,
        activo=1
    ).values('id', 'nombre_producto') 
    lista_productos = list(productos_filtrados)
    
    return JsonResponse({'productos': lista_productos})


def _get_carrito(request, idproveedor):
    carrito_key = f'carrito_compra_{idproveedor}'
    return request.session.get(carrito_key, [])


def _update_carrito(request, idproveedor, carrito):
    carrito_key = f'carrito_compra_{idproveedor}'
    request.session[carrito_key] = carrito
    request.session.modified = True


def _calcular_totales_carrito(carrito):
    subtotal_compra = Decimal('0.00')
    for item in carrito:
        # Convertir de float a Decimal
        subtotal_compra += Decimal(str(item['subtotal']))
    
    iva_compra = subtotal_compra * Decimal('0.19')
    total_compra = subtotal_compra + iva_compra

    return subtotal_compra.quantize(Decimal('0.01')), iva_compra.quantize(Decimal('0.01')), total_compra.quantize(Decimal('0.01'))


@transaction.atomic
@login_required 
def crear_compra_proveedor(request, idproveedor):
    proveedor = get_object_or_404(Proveedor, id=idproveedor)
    categorias = Categoria.objects.filter(activo=1)
    usuario = request.user
    es_admin = request.user.tipo_usu == 'administrador'
    
    print(f"--- DEP-COMPRA: Proveedor ID={proveedor.id}, Nombre={proveedor.nombre_proveedor} ---")

    productos_maestros = Producto.objects.filter(
        activo=1, 
        codigo_barras='SIN_LOTE_CATALOGO', 
        proveedor_idproveedor=proveedor
    ).order_by('nombre_producto')

    stock_consolidado = Producto.objects.filter(
        nombre_producto__in=[p.nombre_producto for p in productos_maestros]
    ).values(
        'nombre_producto'
    ).annotate(
        stock_total=Sum('stock_actual')
    )
    
    stock_dict = {item['nombre_producto']: item['stock_total'] for item in stock_consolidado}

    for p in productos_maestros:
        p.stock_actual = stock_dict.get(p.nombre_producto, 0)
    
    carrito = _get_carrito(request, idproveedor)
    subtotal_compra, iva_compra, total_compra = _calcular_totales_carrito(carrito)
    
    if request.method == "POST":
        
        if 'agregar_producto' in request.POST:
            tipo = request.POST.get("tipo_producto")
            fecha_vencimiento_detalle = request.POST.get("fecha_vencimiento") or None
            lote = request.POST.get("lote", "")
            
            try:
                cantidad = int(request.POST.get("cantidad", "0"))
                precio_unitario = Decimal(request.POST.get("valor_unitario") or "0")
            except:
                messages.error(request, "Error en el formato de la cantidad o el valor unitario.")
                return redirect("crear_compra_proveedor", idproveedor=idproveedor)
        
            if cantidad <= 0 or precio_unitario <= Decimal('0'):
                messages.error(request, "La cantidad y el precio unitario deben ser mayores a cero.")
                return redirect("crear_compra_proveedor", idproveedor=idproveedor)
                
            subtotal_linea = cantidad * precio_unitario

            producto_id = None
            nombre_producto = ""

            if tipo == "nuevo":
                categoria_id = request.POST.get('categoria')
                nombre_producto_post = request.POST.get('nombre_producto')
                
                if not categoria_id or not nombre_producto_post:
                    messages.error(request, "Nombre y Categoría son requeridos para el producto nuevo.")
                    return redirect("crear_compra_proveedor", idproveedor=idproveedor)
                
                categoria_obj = get_object_or_404(Categoria, id=categoria_id)
                
                precio_venta = Decimal('0.00')
                margen_calculado = Decimal('0.00')

                producto = Producto.objects.create(
                    nombre_producto=nombre_producto_post,
                    descripcion_producto=request.POST.get('descripcion_producto', ''),
                    codigo_barras='SIN_LOTE_CATALOGO', 
                    registrosanitario=request.POST.get('registrosanitario', ''),
                    precio_compra=precio_unitario, 
                    precio_venta=precio_venta,
                    margen_ganancia=margen_calculado, 
                    stock_actual=0, 
                    stock_minimo=int(request.POST.get('stock_minimo') or 1),
                    stock_maximo=int(request.POST.get('stock_maximo') or 1000),
                    fecha_vencimiento=date(2099, 12, 31),
                    categoria_idcategoria=categoria_obj, 
                    proveedor_idproveedor_id=proveedor.id,
                    activo=1
                )
                producto_id = producto.id
                nombre_producto = producto.nombre_producto
                
            else:
                producto_id = request.POST.get("producto_id")
                if not producto_id:
                    messages.error(request, "Debe seleccionar un producto existente.")
                    return redirect("crear_compra_proveedor", idproveedor=idproveedor)
                
                producto = get_object_or_404(Producto, id=producto_id)
                nombre_producto = producto.nombre_producto

            item_carrito = {
                'temp_id': str(uuid.uuid4()),
                'producto_id': producto_id,
                'nombre': nombre_producto,
                'cantidad': cantidad,
                'precio': float(precio_unitario),
                'subtotal': float(subtotal_linea.quantize(Decimal('0.01'))),
                'lote': lote,
                'fecha_vencimiento': fecha_vencimiento_detalle,
                'proveedor_id': proveedor.id, 
            }

            carrito.append(item_carrito)
            _update_carrito(request, idproveedor, carrito)

            messages.success(request, f"Producto '{nombre_producto}' añadido a la lista de compra.")
            return redirect("crear_compra_proveedor", idproveedor=idproveedor)

        elif 'finalizar_compra' in request.POST:
            
            if not carrito:
                messages.error(request, "La lista de compra está vacía.")
                return redirect("crear_compra_proveedor", idproveedor=idproveedor)
            
            estado = request.POST.get("estado_compra", "pendiente")
            observaciones = request.POST.get("observaciones", "")
            numero_factura = request.POST.get("numero_factura", "")
            
            subtotal_compra, iva_compra, total_compra = _calcular_totales_carrito(carrito)
            
            compra = Compra_proveedor.objects.create(
                numero_factura_compra=numero_factura if numero_factura else f"CMP{Compra_proveedor.objects.count()+1}",
                subtotal_compra=subtotal_compra,
                iva_compra=iva_compra,
                total_compra=total_compra,
                estado_compra=estado,
                observaciones_compra=observaciones,
                usuarios_id_usuario=request.user,
                proveedor_idproveedor_id=proveedor.id, 
            )

            for item in carrito:
                producto_maestro = get_object_or_404(Producto, id=item['producto_id'])
                
                precio_unitario = Decimal(str(item['precio']))
                fecha_vencimiento_lote = item.get('fecha_vencimiento') or date(2099, 12, 31)
                
                producto_maestro.precio_compra = precio_unitario
                
                if producto_maestro.precio_venta > Decimal('0'):
                    nuevo_margen = ((producto_maestro.precio_venta - precio_unitario) / precio_unitario) * Decimal('100.00')
                    producto_maestro.margen_ganancia = nuevo_margen.quantize(Decimal('0.01'))
                
                producto_maestro.save()
                
                Producto.objects.create(
                    nombre_producto=item['nombre'],
                    descripcion_producto=producto_maestro.descripcion_producto,
                    codigo_barras=item['lote'], 
                    registrosanitario=producto_maestro.registrosanitario,
                    precio_compra=precio_unitario,
                    precio_venta=producto_maestro.precio_venta,
                    margen_ganancia=producto_maestro.margen_ganancia,
                    stock_actual=item['cantidad'] if estado == 'recibida' else 0, 
                    stock_minimo=producto_maestro.stock_minimo,
                    stock_maximo=producto_maestro.stock_maximo,
                    fecha_vencimiento=fecha_vencimiento_lote,
                    categoria_idcategoria=producto_maestro.categoria_idcategoria,
                    proveedor_idproveedor_id=proveedor.id, 
                    activo=1,
                )

                Compra_detalle.objects.create(
                    compra_idcompra=compra,
                    producto_idproducto=producto_maestro, 
                    cantidad=item['cantidad'],
                    precio_compra_unitario=Decimal(str(item['precio'])),
                    subtotal_linea_compra=Decimal(str(item['subtotal'])),
                    lote=item['lote'],
                    fecha_vencimiento=fecha_vencimiento_lote
                )

                if estado == 'recibida':
                    producto_maestro.stock_actual += item['cantidad']
                    producto_maestro.save()
                    
            request.session.pop(f'carrito_compra_{idproveedor}', None)
            messages.success(request, f"Compra #{compra.numero_factura_compra} registrada como '{estado.upper()}'.")
            return redirect("detalle_compra_proveedor", compra_id=compra.id)

        return redirect("crear_compra_proveedor", idproveedor=idproveedor)
        
    return render(request, "proveedor/crear_compra.html", {
        "proveedor": proveedor,
        "productos": productos_maestros, 
        "categorias": categorias,
        "carrito": carrito,
        "subtotal_compra": subtotal_compra,
        "iva_compra": iva_compra,
        "total_compra": total_compra,
        "usuario": usuario,
        "es_admin": es_admin,
    })


@login_required
def compra_quitar_producto(request, idproveedor, temp_id):
    carrito = _get_carrito(request, idproveedor)
    
    carrito_actualizado = [item for item in carrito if item['temp_id'] != temp_id]
    
    if len(carrito_actualizado) < len(carrito):
        _update_carrito(request, idproveedor, carrito_actualizado)
        messages.info(request, "Producto eliminado de la lista de compra.")
    else:
        messages.error(request, "Error: Producto no encontrado en la lista.")

    return redirect("crear_compra_proveedor", idproveedor=idproveedor)


@login_required
def compra_limpiar_carrito(request, idproveedor):
    request.session.pop(f'carrito_compra_{idproveedor}', None)
    request.session.modified = True
    messages.info(request, "Lista de compra limpiada.")
    return redirect("crear_compra_proveedor", idproveedor=idproveedor)


@transaction.atomic
def agregar_al_carrito_compra(request, proveedor_id, productos, categorias):
    proveedor = get_object_or_404(Proveedor, id=proveedor_id)
    
    if request.method != "POST":
        messages.error(request, "Método de solicitud no válido.")
        return redirect('crear_compra_proveedor', idproveedor=proveedor.id)

    tipo = request.POST.get("tipo_producto")

    try:
        cantidad = int(request.POST.get("cantidad", "0"))
        precio_compra_unitario = Decimal(request.POST.get("valor_unitario") or "0") 
        lote_detalle = request.POST.get("lote") 
        fecha_vencimiento = request.POST.get("fecha_vencimiento") or None
        
        descripcion_producto = request.POST.get('descripcion_producto', '').strip() 
        
        if cantidad <= 0 or precio_compra_unitario <= Decimal('0') or not lote_detalle:
            messages.error(request, "Cantidad, Precio Unitario y Lote son obligatorios y deben ser válidos.")
            return redirect('crear_compra_proveedor', idproveedor=proveedor.id)
            
    except Exception as e:
        messages.error(request, f"Error en el formato de datos comunes: {e}")
        return redirect('crear_compra_proveedor', idproveedor=proveedor.id)

    categoria_id_en_carrito = None
    producto_id_en_carrito = 0
    nombre_producto_en_carrito = ""
    precio_venta = Decimal('0.00')

    if tipo == "nuevo":
        nombre_producto = request.POST.get('nombre_producto').strip() 
        categoria_id = request.POST.get('categoria')
        
        precio_venta = Decimal('0.00')
        categoria_id_en_carrito = categoria_id

        if not nombre_producto or not categoria_id:
            messages.error(request, "Faltan datos obligatorios del producto nuevo (Nombre, Categoría).")
            return redirect('crear_compra_proveedor', idproveedor=proveedor.id)
        
        if Producto.objects.filter(
            nombre_producto__iexact=nombre_producto,
            descripcion_producto__iexact=descripcion_producto,
            codigo_barras='SIN_LOTE_CATALOGO'
        ).exists():
             messages.error(request, f"El producto '{nombre_producto}' con esa descripción ya existe en el catálogo. Use la opción 'Existente'.")
             return redirect('crear_compra_proveedor', idproveedor=proveedor.id)
        
        producto_id_en_carrito = 0
        nombre_producto_en_carrito = nombre_producto
    
    elif tipo == "existente":
        producto_id_post = request.POST.get("producto_id")
        producto_existente_obj = get_object_or_404(Producto, id=producto_id_post) 
        
        if producto_existente_obj.codigo_barras != 'SIN_LOTE_CATALOGO':
            messages.error(request, "Error: El ID seleccionado no corresponde a un producto maestro (catálogo).")
            return redirect('crear_compra_proveedor', idproveedor=proveedor.id)
      
        producto_id_en_carrito = producto_existente_obj.id
        nombre_producto_en_carrito = producto_existente_obj.nombre_producto
        categoria_id_en_carrito = producto_existente_obj.categoria_idcategoria.id
        descripcion_producto = producto_existente_obj.descripcion_producto 
        
        precio_venta = producto_existente_obj.precio_venta
        
    else:
        messages.error(request, "Tipo de producto no válido.")
        return redirect('crear_compra_proveedor', idproveedor=proveedor.id)

    carrito_key = f'carrito_compra_{proveedor.id}'
    carrito = request.session.get(carrito_key, [])
    subtotal_linea = cantidad * precio_compra_unitario
    
    nuevo_item = {
        'temp_id': str(uuid4()), 
        'tipo': tipo,
        'producto_id': producto_id_en_carrito,
        'nombre': nombre_producto_en_carrito,
        'precio': float(precio_compra_unitario), 
        'cantidad': cantidad,
        'subtotal': float(subtotal_linea),
        'fecha_vencimiento': fecha_vencimiento,
        'lote': lote_detalle,
        'categoria_id': categoria_id_en_carrito if categoria_id_en_carrito else None,
        'precio_venta_sugerido': float(precio_venta),
        'datos_nuevo_maestro': {
            'registro_sanitario': request.POST.get('registrosaniario', ''),
            'stock_minimo': int(request.POST.get('stock_minimo') or 1),
            'stock_maximo': int(request.POST.get('stock_maximo') or 1000),
            'descripcion_producto': descripcion_producto, 
        }
    }
    
    carrito.append(nuevo_item)
    request.session[carrito_key] = carrito
    request.session.modified = True
    
    messages.success(request, f'Producto añadido a la lista: {nombre_producto_en_carrito}')
    return redirect('crear_compra_proveedor', idproveedor=proveedor.id)


@transaction.atomic
def procesar_compra_final(request, proveedor):
    
    carrito_key = f'carrito_compra_{proveedor.id}'
    carrito = request.session.get(carrito_key, [])
    
    if not carrito:
        messages.error(request, 'La lista de compra está vacía.')
        return redirect('crear_compra_proveedor', idproveedor=proveedor.id) 
        
    estado = request.POST.get("estado_compra", "recibida")
    observaciones = request.POST.get("observaciones", "")
    numero_factura = request.POST.get("numero_factura", "")

    try:
        subtotal_compra = sum(Decimal(str(item['subtotal'])) for item in carrito)
        iva_compra = subtotal_compra * Decimal("0.19") 
        total_compra = subtotal_compra + iva_compra
        
        compra = Compra_proveedor.objects.create(
            numero_factura_compra=numero_factura if numero_factura else f"CMP{Compra_proveedor.objects.count()+1}",
            subtotal_compra=subtotal_compra,
            iva_compra=iva_compra,
            total_compra=total_compra,
            estado_compra=estado,
            observaciones_compra=observaciones,
            imagen_factura_compra="",
            usuarios_id_usuario=request.user,
            proveedor_idproveedor=proveedor
        )
        
        for item in carrito:
            producto_maestro = None
            
            precio_unitario = Decimal(str(item['precio']))
            cantidad = int(item['cantidad'])
            fecha_vencimiento_detalle = item.get('fecha_vencimiento')
            lote_detalle = item['lote']
            subtotal_linea = Decimal(str(item['subtotal']))
            stock_anterior = 0 

            precio_venta_float_sugerido = item.get('precio_venta_sugerido')
            precio_venta_sugerido = Decimal(str(precio_venta_float_sugerido)) if precio_venta_float_sugerido else Decimal('0.00')

            if item['tipo'] == "nuevo":
                datos_maestro = item.get('datos_nuevo_maestro', {})
                fecha_vencimiento_final = fecha_vencimiento_detalle or date(2099, 12, 31)
                categoria_id_from_carrito = item.get('categoria_id')
                
                if not categoria_id_from_carrito:
                    raise Exception("ID de categoría no encontrado en el carrito para producto nuevo.")
                
                categoria_obj = get_object_or_404(Categoria, id=categoria_id_from_carrito)

                precio_venta_maestro = Decimal('0.00')

                descripcion_producto = datos_maestro.get('descripcion_producto', '')
                registrosanitario = datos_maestro.get('registro_sanitario', '')
                stock_minimo_val = datos_maestro.get('stock_minimo') or 1
                stock_maximo_val = datos_maestro.get('stock_maximo') or 999999 

                margen_calculado = Decimal('0.00')
                
                try:
                    producto_maestro, creado_maestro = Producto.objects.get_or_create(
                        nombre_producto=item['nombre'],
                        descripcion_producto=descripcion_producto, 
                        codigo_barras='SIN_LOTE_CATALOGO', 
                        proveedor_idproveedor=proveedor,
                        defaults={
                            'registrosanitario': registrosanitario,
                            'precio_compra': precio_unitario,
                            'precio_venta': precio_venta_maestro,
                            'margen_ganancia': margen_calculado,
                            'stock_actual': 0, 
                            'stock_minimo': stock_minimo_val,
                            'stock_maximo': stock_maximo_val,
                            'fecha_vencimiento': date(2099, 12, 31), 
                            'categoria_idcategoria': categoria_obj,
                            'activo': 1
                        }
                    )
                except IntegrityError as e:
                    transaction.set_rollback(True) 
                    messages.error(request, f"Error de unicidad del catálogo: Ya existe el producto '{item['nombre']}' con la misma descripción.")
                    return redirect('crear_compra_proveedor', idproveedor=proveedor.id)
                
                if not creado_maestro:
                    producto_maestro.precio_compra = precio_unitario
                    if producto_maestro.precio_venta > Decimal('0'):
                        nuevo_margen = ((producto_maestro.precio_venta - precio_unitario) / precio_unitario) * Decimal('100.00')
                        producto_maestro.margen_ganancia = nuevo_margen.quantize(Decimal('0.01'))
                    producto_maestro.save()

                producto = Producto.objects.create(
                    nombre_producto=item['nombre'],
                    descripcion_producto=descripcion_producto, 
                    codigo_barras=lote_detalle, 
                    registrosanitario=registrosanitario, 
                    precio_compra=precio_unitario,
                    precio_venta=producto_maestro.precio_venta,
                    margen_ganancia=producto_maestro.margen_ganancia, 
                    stock_actual=0, 
                    stock_minimo=stock_minimo_val,
                    stock_maximo=stock_maximo_val,
                    fecha_vencimiento=fecha_vencimiento_final,
                    categoria_idcategoria=categoria_obj, 
                    proveedor_idproveedor=proveedor,
                    activo=1
                )
                stock_anterior = 0
            
            else:
                producto_id_from_carrito = item.get('producto_id')
                producto_base = Producto.objects.get(id=producto_id_from_carrito)
                producto_maestro = producto_base 
                
                fecha_vencimiento_final = fecha_vencimiento_detalle or producto_base.fecha_vencimiento or date(2099, 12, 31)

                producto_maestro.precio_compra = precio_unitario
                
                if producto_maestro.precio_venta > Decimal('0'):
                    nuevo_margen = ((producto_maestro.precio_venta - precio_unitario) / precio_unitario) * Decimal('100.00')
                    producto_maestro.margen_ganancia = nuevo_margen.quantize(Decimal('0.01'))
                
                producto_maestro.save()
                
                try:
                    producto = Producto.objects.get(
                        nombre_producto=producto_base.nombre_producto,
                        codigo_barras=lote_detalle, 
                        activo=1
                    )
                    stock_anterior = producto.stock_actual 
                    
                    producto.precio_venta = producto_maestro.precio_venta     
                    producto.margen_ganancia = producto_maestro.margen_ganancia 
                    
                except Producto.DoesNotExist:
                    producto = Producto.objects.create(
                        nombre_producto=producto_base.nombre_producto,
                        descripcion_producto=producto_base.descripcion_producto,
                        categoria_idcategoria=producto_base.categoria_idcategoria,
                        proveedor_idproveedor=proveedor, 
                        codigo_barras=lote_detalle, 
                        fecha_vencimiento=fecha_vencimiento_final,
                        precio_venta=producto_maestro.precio_venta,
                        margen_ganancia=producto_maestro.margen_ganancia, 
                        stock_minimo=producto_base.stock_minimo, 
                        stock_maximo=producto_base.stock_maximo,
                        precio_compra=precio_unitario,
                        stock_actual=0, 
                        registrosanitario=producto_base.registrosanitario,
                        activo=1
                    )
                    stock_anterior = 0 

            if not producto or not producto.pk:
                raise Exception(f"El objeto Producto (Lote) es nulo para '{item['nombre']}'.")
                
            Compra_detalle.objects.create(
                compra_idcompra=compra,
                producto_idproducto=producto, 
                cantidad=cantidad,
                precio_compra_unitario=precio_unitario,
                subtotal_linea_compra=subtotal_linea,
                lote=lote_detalle,
                fecha_vencimiento=fecha_vencimiento_detalle or date(2099, 12, 31)
            )
            
            if estado == 'recibida':
                producto.stock_actual += cantidad
                producto.precio_compra = precio_unitario
                producto.save() 
                
                Movimiento_inventario.objects.create(
                    producto_idproducto=producto, 
                    tipo_movimiento='compra',
                    cantidad=cantidad,
                    stock_anterior=stock_anterior,
                    stock_nuevo=producto.stock_actual,
                    precio_unitario=precio_unitario,
                    valor_total=subtotal_linea,
                    referencia_id=compra.pk,
                    tipo_referencia='compra',
                    observaciones=f"Entrada por compra {compra.numero_factura_compra}. Lote: {lote_detalle}",
                    imagen_comprobante="",
                    usuarios_id_usuario=request.user 
                )

        request.session[carrito_key] = []
        request.session.modified = True
        
        messages.success(request, f"Compra {compra.numero_factura_compra} registrada exitosamente. Total: ${total_compra:,.2f}")
        return redirect("listar_compras") 

    except Exception as e:
        transaction.set_rollback(True) 
        messages.error(request, f'Error CRÍTICO al procesar la compra. Detalles: {str(e)}')
        return redirect('crear_compra_proveedor', idproveedor=proveedor.id)


#DETALLE COMPRA PROVEEDOR
def detalle_compra_proveedor(request, compra_id):

    compra = get_object_or_404(Compra_proveedor, pk=compra_id)  

    detalles = Compra_detalle.objects.filter(compra_idcompra=compra)
    
    proveedor = detalles.first().producto_idproducto.proveedor_idproveedor if detalles.exists() else None

    usuario = request.user
    es_admin = request.user.tipo_usu == 'administrador'


    context = {
        'compra': compra,
        'detalles': detalles,
        'proveedor': proveedor,
        'usuario': usuario,
        'es_admin': es_admin,
    }
    
    return render(request, 'proveedor/detalle_compra_proveedor.html', context)


def listar_compras_proveedor(request):
    compras = Compra_proveedor.objects.all().order_by('-fecha_compra')
    usuario = request.user
    es_admin = request.user.tipo_usu == 'administrador'


    context = {
        'compras': compras,
        'usuario': usuario,
        'es_admin': es_admin,
    }
    
    return render(request, 'proveedor/listar_compras.html', context)


@transaction.atomic
def recibir_compra_pendiente(request, compra_id):
    """
    Cambia el estado de una Compra_proveedor de 'pendiente' a 'recibida'
    y actualiza el stock y movimientos de inventario.
    """
    compra = get_object_or_404(Compra_proveedor, pk=compra_id)
    

    if compra.estado_compra == 'recibida':
        messages.warning(request, f"La compra {compra.numero_factura_compra} ya fue recibida anteriormente.")
        return redirect('detalle_compra_proveedor', compra_id=compra.id)

    detalles = Compra_detalle.objects.filter(compra_idcompra=compra)
    
    if not detalles.exists():
        messages.error(request, f"La compra {compra.numero_factura_compra} no tiene productos asociados.")
        return redirect('detalle_compra_proveedor', compra_id=compra.id)
    
    try:
       
        for detalle in detalles:
            producto = detalle.producto_idproducto
            cantidad = detalle.cantidad
            precio_unitario = detalle.precio_compra_unitario
            subtotal_linea = detalle.subtotal_linea_compra
            lote_detalle = detalle.lote
            
          
            
            stock_anterior = producto.stock_actual
            
         
            producto.stock_actual += cantidad
            producto.precio_compra = precio_unitario 
            
          
            if producto.precio_venta > Decimal('0') and producto.precio_compra > Decimal('0'):
                nuevo_margen = ((producto.precio_venta - producto.precio_compra) / producto.precio_compra) * Decimal('100.00')
                producto.margen_ganancia = nuevo_margen.quantize(Decimal('0.01'))
                
            
            producto.save() 
            
           
            Movimiento_inventario.objects.create(
                producto_idproducto=producto, 
                tipo_movimiento='compra',
                cantidad=cantidad,
                stock_anterior=stock_anterior,
                stock_nuevo=producto.stock_actual,
                precio_unitario=precio_unitario,
                valor_total=subtotal_linea,
                referencia_id=compra.pk,
                tipo_referencia='compra',
                observaciones=f"Entrada por recepción de compra pendiente {compra.numero_factura_compra}. Lote: {lote_detalle}",
                imagen_comprobante="",
                usuarios_id_usuario=request.user 
            )
            
     
        compra.estado_compra = 'recibida'
        compra.save()
        
        messages.success(request, f"La compra {compra.numero_factura_compra} ha sido marcada como 'Recibida' y el inventario actualizado.")
        return redirect('detalle_compra_proveedor', compra_id=compra.id)

    except Exception as e:
        messages.error(request, f'Error al procesar la recepción de la compra {compra.numero_factura_compra}. Detalles: {str(e)}')
        return redirect('detalle_compra_proveedor', compra_id=compra.id)

#PROVEEDOR


@login_required
def listar_proveedores(request):
    usuario = request.user 

    proveedores = Proveedor.objects.all()
    es_admin = usuario.tipo_usu == 'administrador'  
    return render(request, 'proveedor/listar_prov.html', {
        'proveedores': proveedores,
        'usuario': usuario,
        'es_admin': es_admin
    })


@login_required
def registrar_proveedor(request):
    # ⬇️ Se obtiene el usuario logueado para poder mostrar su info en la plantilla
    usuario = request.user  
    
    # ⬇️ Se calcula si el usuario es administrador (para mostrar el badge)
    es_admin = request.user.tipo_usu == 'administrador'

    if request.method == 'POST':
        nombre = request.POST.get('nombre_proveedor')
        correo = request.POST.get('correo_proveedor')
        telefono = request.POST.get('telefono')
        direccion = request.POST.get('direccion')
        nit = request.POST.get('nit')
        contacto_nombre = request.POST.get('contacto_nombre')
        contacto_telefono = request.POST.get('contacto_telefono')
        activo = request.POST.get('activo')
        
        if not nombre or not correo or not telefono or not direccion or not nit:
            messages.error(request, "Todos los campos obligatorios deben llenarse.")
            return redirect('registrar_proveedor')
        
        proveedor = Proveedor(
            nombre_proveedor=nombre,
            correo_proveedor=correo,
            telefono=telefono,
            direccion=direccion,
            nit=nit,
            contacto_nombre=contacto_nombre,
            contacto_telefono=contacto_telefono,
            activo=1 if activo == 'True' else 0
        )
        proveedor.save()

        return redirect('listar_proveedores')

    # ⬇️ Se envían las variables al contexto para que la plantilla pueda mostrar el badge
    return render(request, 'proveedor/registrarprov.html', {
        'usuario': usuario,
        'es_admin': es_admin
    })


@login_required
def editar_proveedor(request, id):
    proveedor = get_object_or_404(Proveedor, id=id)

    if request.method == 'POST':
        proveedor.nombre_proveedor = request.POST['nombre_proveedor']
        proveedor.correo_proveedor = request.POST['correo_proveedor']
        proveedor.telefono = request.POST['telefono']
        proveedor.direccion = request.POST['direccion']
        proveedor.nit = request.POST['nit']
        proveedor.contacto_nombre = request.POST['contacto_nombre']
        proveedor.contacto_telefono = request.POST['contacto_telefono']
        proveedor.activo = 1 if request.POST.get('activo') == 'True' else 0
        proveedor.save()
        return redirect('listar_proveedores')

    usuario = request.user                     # <-- usuario logueado
    es_admin = request.user.tipo_usu == 'administrador'  # <-- bandera del rol

    return render(request, 'proveedor/editar_proveedor.html', {
        'proveedor': proveedor,
        'usuario': usuario,      
        'es_admin': es_admin     
    })


@login_required
def eliminar_proveedor(request, id):
    proveedor = get_object_or_404(Proveedor, id=id)
    proveedor.delete()
    return redirect('listar_proveedores')

@login_required(login_url='login')
def proveedores_generar_pdf(request):
    """
    Genera un PDF con el listado completo de proveedores.
    """
    from reportlab.lib.pagesizes import letter
    from reportlab.lib import colors
    from reportlab.lib.units import inch
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.enums import TA_CENTER
    from django.http import HttpResponse
    from io import BytesIO
    from datetime import datetime

    # Obtener proveedores
    proveedores = Proveedor.objects.all().order_by('nombre_proveedor')

    # Preparar respuesta
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="Listado_Proveedores.pdf"'

    # Crear PDF
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=letter)
    elements = []

    # Estilos
    styles = getSampleStyleSheet()
    styleN = styles["Normal"]
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=24,
        textColor=colors.HexColor('#3d862e'),
        alignment=TA_CENTER
    )

    # TÍTULO
    elements.append(Paragraph("ROMAR NATURAL", title_style))

    elements.append(Spacer(1, 0.3 * inch))

    # SUBTÍTULO
    elements.append(Paragraph("<b>LISTADO DE PROVEEDORES</b>", styles['Heading2']))

    # FECHA DE GENERACIÓN
    fecha = datetime.now().strftime('%d/%m/%Y %H:%M')
    elements.append(Paragraph(f"Generado el: {fecha}", styles['Normal']))

    elements.append(Spacer(1, 0.3 * inch))

    # TABLA DE PROVEEDORES
    data = [
        [
            'ID',
            'Nombre',
            'Correo',
            'Teléfono',
            'Dirección',
            'NIT',
            'Contacto',
            'Tel.Contacto',
            'Activo'
        ]
    ]

    for idx, p in enumerate(proveedores, 1):
        data.append([
            str(idx),
            p.nombre_proveedor[:25],
            p.correo_proveedor[:25],
            p.telefono,
            p.direccion[:25],
            p.nit,
            p.contacto_nombre if p.contacto_nombre else "",
            p.contacto_telefono if p.contacto_telefono else "",
            "Sí" if p.activo == 1 else "No",
        ])

    # Definir columnas
    # Calcular ancho disponible (página - márgenes por defecto)
    page_width, page_height = letter
    left_margin = doc.leftMargin
    right_margin = doc.rightMargin
    disponible = page_width - left_margin - right_margin

    # Pesos proporcionales (no son anchos, son "relaciones")
    pesos = [0.5, 1.5, 1.5, 1, 1.7, 1, 1.4, 1.1, 0.7]

    total_pesos = sum(pesos)

    # Convertir pesos → anchos reales en pulgadas
    col_widths = [(p / total_pesos) * disponible for p in pesos]

    table = Table(data, colWidths=col_widths)

    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#3d862e')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 11),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 10),

        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('FONTSIZE', (0, 1), (-1, -1), 9),
    ]))

    elements.append(table)

    # Construir PDF
    doc.build(elements)
    pdf = buffer.getvalue()
    buffer.close()
    response.write(pdf)

    return response



#LOGIN - AUTENTICACIÓN 

def login_view(request):
    """
    Vista para login de usuarios - SIN usar django.contrib.auth.login()
    """
    if request.user.is_authenticated:
        return redirect('dashboard')
    
    login_ok = False  # <-- FLAG para SweetAlert

    if request.method == 'POST':
        form = LoginForm(request.POST)
        if form.is_valid():
            correo = form.cleaned_data['correo']
            password = form.cleaned_data['password']
            
            # Autenticar usando el backend personalizado
            user = authenticate(request, correo=correo, password=password)
            
            if user is not None:
                # Crear sesión manualmente sin disparar señales
                request.session['_auth_user_id'] = user.pk
                request.session['_auth_user_backend'] = 'Sgiev.backends.UsuariosBackend'
                request.session.save()
                
                # ACTIVAMOS EL SWEET ALERT (no redirigimos aquí)
                login_ok = True
            else:
                messages.error(request, 'Credenciales inválidas')
    else:
        form = LoginForm()
    
    return render(request, 'login.html', {'form': form, 'login_ok': login_ok})



def logout_view(request):
    """
    Vista para cerrar sesión
    """
    logout(request)
    messages.success(request, 'Sesión cerrada exitosamente')
    return redirect('index')


# ===== VISTAS DE USUARIOS (CRUD) =====

@admin_required
def usuarios_listar(request):
    """
    Lista todos los usuarios con paginación y búsqueda
    """
    # Obtener parámetro de búsqueda
    search = request.GET.get('search', '')
    
    # Filtrar usuarios
    if search:
        usuarios = Usuarios.objects.filter(
            models.Q(p_nombre__icontains=search) |
            models.Q(p_apellido__icontains=search) |
            models.Q(correo__icontains=search) |
            models.Q(num_identificacion__icontains=search)
        ).order_by('-fecha_registro')
    else:
        usuarios = Usuarios.objects.all().order_by('-fecha_registro')
    
    # Paginación
    paginator = Paginator(usuarios, 10)  # 10 usuarios por página
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    context = {
        'page_obj': page_obj,
        'search': search,
        'usuario': request.user
    }
    
    return render(request, 'usuarios/listar.html', context)


from django.core.mail import send_mail
from django.conf import settings

@admin_required
def usuarios_crear(request):
    """
    Crear un nuevo usuario y enviarle notificación por correo.
    """
    if request.method == 'POST':
        form = UsuarioForm(request.POST)
        if form.is_valid():

            usuario = form.save()  # guardar usuario

            # Enviar correo al nuevo usuario (si tiene correo)
            if usuario.correo:
                try:
                    send_mail(
                        subject='Tu usuario ha sido creado - Romar Natural',
                        message=(
                            f'Hola {usuario.nombre_completo},\n\n'
                            f'Tu usuario ha sido creado en el sistema Romar Natural.\n'
                            f'Correo registrado: {usuario.correo}\n\n'
                            f'Si no esperabas este mensaje, contacta con el administrador.'
                        ),
                        from_email=settings.DEFAULT_FROM_EMAIL,
                        recipient_list=[usuario.correo],
                        fail_silently=False,
                    )
                except Exception as e:
                    print("ERROR EMAIL USUARIO:", e)
                    messages.error(request, f'No se pudo enviar el correo: {e}')

            messages.success(request, 'Usuario creado exitosamente')
            return redirect('usuarios_listar')
    else:
        form = UsuarioForm()

    context = {
        'form': form,
        'titulo': 'Crear Usuario',
        'usuario': request.user,
    }

    return render(request, 'usuarios/crear.html', context)

@admin_required
def usuarios_editar(request, id):
    """
    Editar un usuario existente
    """
    usuario = get_object_or_404(Usuarios, pk=id)
    
    if request.method == 'POST':
        form = UsuarioForm(request.POST, instance=usuario)
        if form.is_valid():
            form.save()
            messages.success(request, 'Usuario actualizado exitosamente')
            return redirect('usuarios_listar')
    else:
        form = UsuarioForm(instance=usuario)
    
    context = {
        'form': form,
        'titulo': 'Editar Usuario',
        'usuario': request.user,
        'usuario_editando': usuario
    }
    
    return render(request, 'usuarios/editar.html', context)


@admin_required
def usuarios_eliminar(request, id):
    """
    Eliminar (desactivar) un usuario
    """
    usuario = get_object_or_404(Usuarios, pk=id)
    
    # No permitir eliminar al usuario actual
    if usuario.id == request.user.id:
        messages.error(request, 'No puedes eliminar tu propio usuario')
        return redirect('usuarios_listar')
    
    # Desactivar en lugar de eliminar
    usuario.activo = 0
    usuario.save()
    
    messages.success(request, f'Usuario {usuario.nombre_completo} desactivado exitosamente')
    return redirect('usuarios_listar')


@admin_required
def usuarios_detalle(request, id):
    """
    Ver detalles de un usuario
    """
    usuario_detalle = get_object_or_404(Usuarios, pk=id)
    
    context = {
        'usuario_detalle': usuario_detalle,
        'usuario': request.user
    }
    
    return render(request, 'usuarios/detalle.html', context)

@login_required
def perfil_usuario(request):
    usuario_detalle = request.user  # ← El usuario logeado

    context = {
        'usuario_detalle': usuario_detalle,
        'usuario': request.user
    }

    return render(request, 'usuarios/detalle_usu.html', context)






# ===== VISTAS DE VENTAS =====
@login_required(login_url='login')
def ventas_listar(request):
    """
    Lista todas las ventas con filtros
    - Administrador: ve todas las ventas
    - Operario: solo ve sus propias ventas
    """
    # Filtros
    search = request.GET.get('search', '')
    estado = request.GET.get('estado', '')
    metodo = request.GET.get('metodo', '')
    
    # Query base según el rol
    if request.user.tipo_usu == 'administrador':
        # Admin ve todas las ventas
        ventas = Venta.objects.all().select_related('usuarios_id_usuario').order_by('-fecha_factura')
    else:
        # Operario solo ve sus ventas
        ventas = Venta.objects.filter(usuarios_id_usuario=request.user).select_related('usuarios_id_usuario').order_by('-fecha_factura')
    
    # Aplicar filtros
    if search:
        ventas = ventas.filter(
            models.Q(numero_factura__icontains=search) |
            models.Q(usuarios_id_usuario__p_nombre__icontains=search) |
            models.Q(usuarios_id_usuario__p_apellido__icontains=search)
        )
    
    if estado:
        ventas = ventas.filter(estado_pago=estado)
    
    if metodo:
        ventas = ventas.filter(metodo_pago=metodo)
    
    # Paginación
    paginator = Paginator(ventas, 10)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    # ===== AGREGAR INFORMACIÓN DE ENVÍOS PARA CADA VENTA =====
    # Crear un diccionario con los envíos asociados a cada venta
    ventas_ids = [venta.id for venta in page_obj]
    envios_dict = {}
    
    if ventas_ids:
        envios = Envio.objects.filter(venta_idfactura__in=ventas_ids).select_related('venta_idfactura')
        for envio in envios:
            envios_dict[envio.venta_idfactura.id] = envio
    
    # Agregar el envío a cada venta en el page_obj
    for venta in page_obj:
        venta.envio_asociado = envios_dict.get(venta.id, None)
    
    context = {
        'page_obj': page_obj,
        'search': search,
        'estado': estado,
        'metodo': metodo,
        'usuario': request.user,
        'es_admin': request.user.tipo_usu == 'administrador'
    }
    
    return render(request, 'ventas/listar.html', context)

@login_required(login_url='login')
def ventas_crear(request):
    """
    Crear nueva venta con sistema de carrito.
    Ahora también captura datos del cliente.
    """
    if request.method == 'POST':
        if 'finalizar_venta' in request.POST:
            # Finalizar venta
            return procesar_venta(request)
        else:
            # Agregar producto al carrito
            return agregar_al_carrito(request)

    # GET - Mostrar formulario
    venta_form = VentaForm()
    producto_form = AgregarProductoForm()

    # Obtener carrito de la sesión
    carrito = request.session.get('carrito_venta', [])

    # Calcular totales
    subtotal = sum(Decimal(str(item['subtotal'])) for item in carrito)

    context = {
        'venta_form': venta_form,
        'producto_form': producto_form,
        'carrito': carrito,
        'subtotal': subtotal,
        'usuario': request.user,
        'es_admin': request.user.tipo_usu == 'administrador',
    }

    return render(request, 'ventas/crear.html', context)



def agregar_al_carrito(request):
    """
    Agrega un producto al carrito de la sesión
    """
    producto_form = AgregarProductoForm(request.POST)

    if producto_form.is_valid():
        producto = producto_form.cleaned_data['producto']
        cantidad = producto_form.cleaned_data['cantidad']

        # Obtener o crear carrito
        carrito = request.session.get('carrito_venta', [])

        # Verificar si el producto ya está en el carrito
        producto_existente = False
        for item in carrito:
            if item['producto_id'] == producto.id:
                # Actualizar cantidad
                nueva_cantidad = item['cantidad'] + cantidad
                if nueva_cantidad <= producto.stock_actual:
                    item['cantidad'] = nueva_cantidad
                    item['subtotal'] = float(producto.precio_venta * nueva_cantidad)
                    item['stock_disponible'] = producto.stock_actual  # Actualizar stock
                    producto_existente = True
                    messages.success(request, f'Cantidad actualizada: {producto.nombre_producto}')
                else:
                    messages.error(request, f'Stock insuficiente para {producto.nombre_producto}')
                    return redirect('ventas_crear')
                break

        if not producto_existente:
            # Agregar nuevo producto
            carrito.append({
                'producto_id': producto.id,
                'nombre': producto.nombre_producto,
                'precio': float(producto.precio_venta),
                'cantidad': cantidad,
                'subtotal': float(producto.precio_venta * cantidad),
                'stock_disponible': producto.stock_actual,
                'stock_minimo': producto.stock_minimo,
            })
            messages.success(request, f'Producto agregado: {producto.nombre_producto}')

        # Guardar carrito en sesión
        request.session['carrito_venta'] = carrito
        request.session.modified = True
    else:
        for error in producto_form.errors.values():
            messages.error(request, error)

    return redirect('ventas_crear')



@login_required(login_url='login')
def ventas_quitar_producto(request, producto_id):
    """
    Quita un producto del carrito
    """
    carrito = request.session.get('carrito_venta', [])
    carrito = [item for item in carrito if item['producto_id'] != producto_id]
    request.session['carrito_venta'] = carrito
    request.session.modified = True
    
    messages.success(request, 'Producto eliminado del carrito')
    return redirect('ventas_crear')


@login_required(login_url='login')
def ventas_limpiar_carrito(request):
    """
    Limpia todo el carrito
    """
    request.session['carrito_venta'] = []
    request.session.modified = True
    
    messages.success(request, 'Carrito vaciado')
    return redirect('ventas_crear')


@transaction.atomic
def procesar_venta(request):
    """
    Procesa la venta final con validación de abono mínimo y estado automático,
    descontando stock por lotes (FIFO) a partir del producto maestro.
    Incluye datos de cliente y envío de correos con PDF adjunto.
    """
    from django.core.mail import EmailMessage
    from django.conf import settings
    from reportlab.lib.pagesizes import letter
    from reportlab.lib import colors
    from reportlab.lib.units import inch
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.enums import TA_CENTER
    from io import BytesIO
    
    venta_form = VentaForm(request.POST)
    carrito = request.session.get('carrito_venta', [])

    if not carrito:
        messages.error(request, 'El carrito está vacío')
        return redirect('ventas_crear')

    if venta_form.is_valid():
        try:
            # ===== CALCULAR TOTALES =====
            subtotal = sum(Decimal(str(item['subtotal'])) for item in carrito)
            descuento = venta_form.cleaned_data['descuento'] or Decimal('0')
            iva = (subtotal - descuento) * Decimal('0.19')
            valor_total = subtotal - descuento + iva
            abono = venta_form.cleaned_data['abono'] or Decimal('0')

            # ===== VALIDACIÓN DE ABONO MÍNIMO =====
            abono_minimo = valor_total * Decimal('0.10')

            if abono > 0 and abono < abono_minimo:
                messages.error(
                    request,
                    f'El abono mínimo debe ser el 10% del total (${abono_minimo:,.0f}). '
                    f'Si no desea abonar, deje el campo en 0.'
                )
                return redirect('ventas_crear')

            # ===== CALCULAR ESTADO DE PAGO AUTOMÁTICAMENTE =====
            if abono == 0:
                estado_pago = 'pendiente'
            elif abono >= valor_total:
                estado_pago = 'pagado'
                abono = valor_total
            else:
                estado_pago = 'parcial'

            saldo_pendiente = valor_total - abono

            # ===== CREAR VENTA =====
            venta = venta_form.save(commit=False)
            venta.subtotal = subtotal
            venta.iva = iva
            venta.valor_total = valor_total
            venta.abono = abono
            venta.saldo_pendiente = saldo_pendiente
            venta.estado_pago = estado_pago
            venta.usuarios_id_usuario = request.user
            venta.imagen_recibo = ''
            venta.save()

            productos_stock_bajo = []

            # ===== PROCESAR CADA PRODUCTO DEL CARRITO (FIFO POR LOTES) =====
            for item in carrito:
                producto_maestro = Producto.objects.get(id=item['producto_id'])
                cantidad_a_vender = item['cantidad']

                lotes = Producto.objects.filter(
                    nombre_producto=producto_maestro.nombre_producto,
                    descripcion_producto=producto_maestro.descripcion_producto,
                    activo=1
                ).order_by('fecha_vencimiento', 'id')

                stock_total_disponible = sum(lote.stock_actual for lote in lotes)
                if stock_total_disponible < cantidad_a_vender:
                    raise Exception(f'Stock insuficiente para {producto_maestro.nombre_producto}')

                Venta_has_producto.objects.create(
                    venta_idfactura=venta,
                    producto_idproducto=producto_maestro,
                    cantidad=cantidad_a_vender,
                    valor_unitario=producto_maestro.precio_venta,
                    subtotal_linea=Decimal(str(item['subtotal']))
                )

                restante = cantidad_a_vender
                for lote in lotes:
                    if restante <= 0:
                        break

                    tomar = min(restante, lote.stock_actual)
                    if tomar <= 0:
                        continue

                    stock_anterior = lote.stock_actual
                    lote.stock_actual -= tomar
                    lote.save()

                    if lote.stock_actual <= lote.stock_minimo:
                        productos_stock_bajo.append({
                            'nombre': lote.nombre_producto,
                            'stock_actual': lote.stock_actual,
                            'stock_minimo': lote.stock_minimo
                        })

                    Movimiento_inventario.objects.create(
                        tipo_movimiento='venta',
                        cantidad=tomar,
                        stock_anterior=stock_anterior,
                        stock_nuevo=lote.stock_actual,
                        precio_unitario=producto_maestro.precio_venta,
                        valor_total=producto_maestro.precio_venta * tomar,
                        referencia_id=venta.id,
                        tipo_referencia='venta',
                        observaciones=f'Venta #{venta.numero_factura} - Lote {lote.codigo_barras}',
                        imagen_comprobante='',
                        usuarios_id_usuario=request.user,
                        producto_idproducto=lote
                    )

                    restante -= tomar

            # ===== LIMPIAR CARRITO =====
            request.session['carrito_venta'] = []
            request.session.modified = True

            # ===== GENERAR PDF EN MEMORIA =====
            productos_venta = Venta_has_producto.objects.filter(venta_idfactura=venta)
            
            buffer = BytesIO()
            doc = SimpleDocTemplate(buffer, pagesize=letter)
            elements = []

            styles = getSampleStyleSheet()
            title_style = ParagraphStyle(
                'CustomTitle',
                parent=styles['Heading1'],
                fontSize=24,
                textColor=colors.HexColor('#3d862e'),
                alignment=TA_CENTER
            )

            elements.append(Paragraph("ROMAR NATURAL", title_style))
            elements.append(Paragraph("NIT: 52101085", styles['Normal']))
            elements.append(Paragraph("Teléfono: 3053615676", styles['Normal']))
            elements.append(Spacer(1, 0.3 * inch))

            elements.append(Paragraph(f"<b>FACTURA: {venta.numero_factura}</b>", styles['Heading2']))
            elements.append(Paragraph(f"Fecha: {venta.fecha_factura.strftime('%d/%m/%Y %H:%M')}", styles['Normal']))
            elements.append(Paragraph(f"Vendedor: {venta.usuarios_id_usuario.nombre_completo}", styles['Normal']))

            if venta.nombre_cliente or venta.correo_cliente or venta.direccion_cliente:
                elements.append(Spacer(1, 0.2 * inch))
                elements.append(Paragraph("<b>Datos del Cliente</b>", styles['Heading3']))
                if venta.nombre_cliente:
                    elements.append(Paragraph(f"Nombre: {venta.nombre_cliente}", styles['Normal']))
                if venta.correo_cliente:
                    elements.append(Paragraph(f"Correo: {venta.correo_cliente}", styles['Normal']))
                if venta.telefono_cliente:
                    elements.append(Paragraph(f"Teléfono: {venta.telefono_cliente}", styles['Normal']))
                if venta.direccion_cliente:
                    elements.append(Paragraph(f"Dirección: {venta.direccion_cliente}", styles['Normal']))

            elements.append(Spacer(1, 0.3 * inch))

            data = [['#', 'Producto', 'Cant.', 'Precio Unit.', 'Subtotal']]
            for idx, item in enumerate(productos_venta, 1):
                data.append([
                    str(idx),
                    item.producto_idproducto.nombre_producto[:30],
                    str(item.cantidad),
                    f"${item.valor_unitario:,.0f}",
                    f"${item.subtotal_linea:,.0f}",
                ])

            table = Table(data, colWidths=[0.5 * inch, 3 * inch, 0.8 * inch, 1.2 * inch, 1.2 * inch])
            table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#3d862e')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 12),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ]))

            elements.append(table)
            elements.append(Spacer(1, 0.3 * inch))

            totales_data = [
                ['Subtotal:', f"${venta.subtotal:,.0f}"],
                ['Descuento:', f"-${venta.descuento:,.0f}"],
                ['IVA (19%):', f"${venta.iva:,.0f}"],
                ['<b>TOTAL:</b>', f"<b>${venta.valor_total:,.0f}</b>"],
            ]

            if venta.abono > 0:
                totales_data.append(['Abono:', f"${venta.abono:,.0f}"])
                totales_data.append(['<b>Saldo Pendiente:</b>', f"<b>${venta.saldo_pendiente:,.0f}</b>"])

            totales_table = Table(totales_data, colWidths=[3 * inch, 2 * inch])
            totales_table.setStyle(TableStyle([
                ('ALIGN', (0, 0), (-1, -1), 'RIGHT'),
                ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
                ('FONTSIZE', (0, -1), (-1, -1), 14),
                ('LINEABOVE', (0, -2), (-1, -2), 2, colors.black),
            ]))

            elements.append(totales_table)
            elements.append(Spacer(1, 0.5 * inch))

            elements.append(Paragraph(f"<b>Método de Pago:</b> {venta.get_metodo_pago_display()}", styles['Normal']))
            elements.append(Paragraph(f"<b>Estado:</b> {venta.get_estado_pago_display()}", styles['Normal']))

            if venta.observaciones:
                elements.append(Spacer(1, 0.2 * inch))
                elements.append(Paragraph(f"<b>Observaciones:</b> {venta.observaciones}", styles['Normal']))

            elements.append(Spacer(1, 0.5 * inch))
            elements.append(Paragraph("Gracias por su compra", styles['Normal']))

            doc.build(elements)
            pdf_content = buffer.getvalue()
            buffer.close()

            # ===== ENVIAR CORREO AL CLIENTE =====
            if venta.correo_cliente:
                try:
                    email_cliente = EmailMessage(
                        subject=f'Venta registrada - {venta.numero_factura}',
                        body=(
                            f'Hola {venta.nombre_cliente},\n\n'
                            f'Tu compra ha sido registrada exitosamente.\n'
                            f'Factura: {venta.numero_factura}\n'
                            f'Total: ${venta.valor_total:,.0f}\n\n'
                            f'Pronto tu pedido entrará en proceso de envío.\n'
                            f'Adjuntamos tu factura en formato PDF.\n\n'
                            f'Gracias por tu compra.\n\n'
                            f'Romar Natural'
                        ),
                        from_email=settings.DEFAULT_FROM_EMAIL,
                        to=[venta.correo_cliente],
                    )
                    email_cliente.attach(f'Factura_{venta.numero_factura}.pdf', pdf_content, 'application/pdf')
                    email_cliente.send(fail_silently=True)
                except Exception:
                    pass

            # ===== ENVIAR CORREO AL VENDEDOR =====
            if venta.usuarios_id_usuario.correo:
                try:
                    email_vendedor = EmailMessage(
                        subject=f'Venta registrada exitosamente - {venta.numero_factura}',
                        body=(
                            f'Hola {venta.usuarios_id_usuario.nombre_completo},\n\n'
                            f'La venta {venta.numero_factura} ha sido registrada exitosamente.\n'
                            f'Cliente: {venta.nombre_cliente or "N/A"}\n'
                            f'Total: ${venta.valor_total:,.0f}\n\n'
                            f'Por favor, realiza la asignación del envío lo más pronto posible.\n'
                            f'Adjuntamos la factura en formato PDF.\n\n'
                            f'Romar Natural'
                        ),
                        from_email=settings.DEFAULT_FROM_EMAIL,
                        to=[venta.usuarios_id_usuario.correo],
                    )
                    email_vendedor.attach(f'Factura_{venta.numero_factura}.pdf', pdf_content, 'application/pdf')
                    email_vendedor.send(fail_silently=True)
                except Exception:
                    pass

            # ===== MENSAJES DE ÉXITO =====
            estado_texto = {
                'pendiente': 'PENDIENTE',
                'parcial': 'PARCIAL',
                'pagado': 'PAGADO'
            }
            messages.success(
                request,
                f'Venta {venta.numero_factura} registrada exitosamente. '
                f'Estado: {estado_texto[estado_pago]}'
            )

            if productos_stock_bajo:
                for prod in productos_stock_bajo:
                    messages.warning(
                        request,
                        f'⚠️ ALERTA: {prod["nombre"]} tiene stock bajo. '
                        f'Actual: {prod["stock_actual"]} | Mínimo: {prod["stock_minimo"]}'
                    )

            return redirect('ventas_detalle', id=venta.id)

        except Exception as e:
            messages.error(request, f'Error al procesar la venta: {str(e)}')
            return redirect('ventas_crear')
    else:
        for error in venta_form.errors.values():
            messages.error(request, error)
        return redirect('ventas_crear')

@login_required(login_url='login')
def ventas_detalle(request, id):
    """
    Ver detalles completos de una venta (factura),
    incluyendo datos del cliente.
    """
    venta = get_object_or_404(Venta, pk=id)
    productos = Venta_has_producto.objects.filter(venta_idfactura=venta)

    context = {
        'venta': venta,
        'productos': productos,
        'usuario': request.user,
        'es_admin': request.user.tipo_usu == 'administrador',
         }

    return render(request, 'ventas/detalle.html', context)


@login_required(login_url='login')
def obtener_precio_producto(request, producto_id):
    """
    API para obtener información de un producto (AJAX)
    """
    try:
        producto = Producto.objects.get(id=producto_id, activo=1)
        data = {
            'precio': float(producto.precio_venta),
            'stock': producto.stock_actual,
            'nombre': producto.nombre_producto
        }
        return JsonResponse(data)
    except Producto.DoesNotExist:
        return JsonResponse({'error': 'Producto no encontrado'}, status=404)

@admin_required
@transaction.atomic
def ventas_editar_estado(request, id):
    """
    Permite al administrador editar el estado de pago de una venta,
    agregar abonos adicionales y actualizar información del cliente.
    """
    venta = get_object_or_404(Venta, pk=id)

    if request.method == 'POST':
        form = EditarEstadoVentaForm(request.POST, instance=venta, venta=venta)
        
        if form.is_valid():
            nuevo_abono = form.cleaned_data.get('nuevo_abono') or Decimal('0')
            
            # Guardar los datos del cliente y observaciones
            venta = form.save(commit=False)
            
            # Procesar nuevo abono si existe
            if nuevo_abono > 0:
                # Actualizar abono total
                venta.abono += nuevo_abono
                
                # Recalcular saldo pendiente
                venta.saldo_pendiente = venta.valor_total - venta.abono
                
                # Actualizar estado automáticamente
                if venta.abono >= venta.valor_total:
                    venta.estado_pago = 'pagado'
                    venta.saldo_pendiente = Decimal('0')
                elif venta.abono > 0:
                    venta.estado_pago = 'parcial'
                else:
                    venta.estado_pago = 'pendiente'
                
                messages.success(
                    request,
                    f'Abono de ${nuevo_abono:,.0f} registrado exitosamente. '
                    f'Nuevo estado: {venta.get_estado_pago_display()}'
                )
            
            venta.save()
            
            messages.success(
                request,
                f'Información de la venta {venta.numero_factura} actualizada correctamente'
            )
            return redirect('ventas_detalle', id=venta.id)
    else:
        form = EditarEstadoVentaForm(instance=venta, venta=venta)

    context = {
        'form': form,
        'venta': venta,
        'usuario': request.user,
        'es_admin': True,
    }

    return render(request, 'ventas/editar_estado.html', context)

@login_required(login_url='login')
def ventas_generar_pdf(request, id):
    """
    Genera un PDF de la factura de venta, incluyendo datos del cliente.
    """
    from reportlab.lib.pagesizes import letter
    from reportlab.lib import colors
    from reportlab.lib.units import inch
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.enums import TA_CENTER, TA_RIGHT
    from django.http import HttpResponse
    from io import BytesIO
    # Obtener la venta
    venta = get_object_or_404(Venta, pk=id)
    productos = Venta_has_producto.objects.filter(venta_idfactura=venta)

    # Crear el objeto HttpResponse con el tipo de contenido PDF
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="Factura_{venta.numero_factura}.pdf"'

    # Crear el PDF
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=letter)
    elements = []

    # Estilos
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=24,
        textColor=colors.HexColor('#3d862e'),
        alignment=TA_CENTER
    )

    # Título
    elements.append(Paragraph("ROMAR NATURAL", title_style))
    elements.append(Paragraph("NIT: 52101085", styles['Normal']))
    elements.append(Paragraph("Teléfono: 3053615676", styles['Normal']))
    elements.append(Spacer(1, 0.3 * inch))

    # Información de la factura
    elements.append(Paragraph(f"<b>FACTURA: {venta.numero_factura}</b>", styles['Heading2']))
    elements.append(Paragraph(f"Fecha: {venta.fecha_factura.strftime('%d/%m/%Y %H:%M')}", styles['Normal']))
    elements.append(Paragraph(f"Vendedor: {venta.usuarios_id_usuario.nombre_completo}", styles['Normal']))

    # Información del cliente
    if venta.nombre_cliente or venta.correo_cliente or venta.direccion_cliente:
        elements.append(Spacer(1, 0.2 * inch))
        elements.append(Paragraph("<b>Datos del Cliente</b>", styles['Heading3']))
        if venta.nombre_cliente:
            elements.append(Paragraph(f"Nombre: {venta.nombre_cliente}", styles['Normal']))
        if venta.correo_cliente:
            elements.append(Paragraph(f"Correo: {venta.correo_cliente}", styles['Normal']))
        if venta.telefono_cliente:
            elements.append(Paragraph(f"Teléfono: {venta.telefono_cliente}", styles['Normal']))
        if venta.direccion_cliente:
            elements.append(Paragraph(f"Dirección: {venta.direccion_cliente}", styles['Normal']))

    elements.append(Spacer(1, 0.3 * inch))

    # Tabla de productos
    data = [['#', 'Producto', 'Cant.', 'Precio Unit.', 'Subtotal']]

    for idx, item in enumerate(productos, 1):
        data.append([
            str(idx),
            item.producto_idproducto.nombre_producto[:30],
            str(item.cantidad),
            f"${item.valor_unitario:,.0f}",
            f"${item.subtotal_linea:,.0f}",
        ])

    table = Table(data, colWidths=[0.5 * inch, 3 * inch, 0.8 * inch, 1.2 * inch, 1.2 * inch])
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#3d862e')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 12),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
    ]))

    elements.append(table)
    elements.append(Spacer(1, 0.3 * inch))

    # Totales
    totales_data = [
        ['Subtotal:', f"${venta.subtotal:,.0f}"],
        ['Descuento:', f"-${venta.descuento:,.0f}"],
        ['IVA (19%):', f"${venta.iva:,.0f}"],
        ['<b>TOTAL:</b>', f"<b>${venta.valor_total:,.0f}</b>"],
    ]

    if venta.abono > 0:
        totales_data.append(['Abono:', f"${venta.abono:,.0f}"])
        totales_data.append(['<b>Saldo Pendiente:</b>', f"<b>${venta.saldo_pendiente:,.0f}</b>"])

    totales_table = Table(totales_data, colWidths=[3 * inch, 2 * inch])
    totales_table.setStyle(TableStyle([
        ('ALIGN', (0, 0), (-1, -1), 'RIGHT'),
        ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
        ('FONTSIZE', (0, -1), (-1, -1), 14),
        ('LINEABOVE', (0, -2), (-1, -2), 2, colors.black),
    ]))

    elements.append(totales_table)
    elements.append(Spacer(1, 0.5 * inch))

    # Información adicional
    elements.append(Paragraph(f"<b>Método de Pago:</b> {venta.get_metodo_pago_display()}", styles['Normal']))
    elements.append(Paragraph(f"<b>Estado:</b> {venta.get_estado_pago_display()}", styles['Normal']))

    if venta.observaciones:
        elements.append(Spacer(1, 0.2 * inch))
        elements.append(Paragraph(f"<b>Observaciones:</b> {venta.observaciones}", styles['Normal']))

    elements.append(Spacer(1, 0.5 * inch))
    elements.append(Paragraph("Gracias por su compra", styles['Normal']))

    # Construir PDF
    doc.build(elements)

    pdf = buffer.getvalue()
    buffer.close()
    response.write(pdf)

    return response

@admin_required
@transaction.atomic
def ventas_eliminar(request, id):
    """
    Elimina una venta y devuelve el stock al inventario.
    Solo disponible para administradores.
    """
    venta = get_object_or_404(Venta, pk=id)

    try:
        # Obtener todos los productos de la venta
        productos_venta = Venta_has_producto.objects.filter(venta_idfactura=venta)

        # Revertir el stock de cada producto
        for item in productos_venta:
            producto = item.producto_idproducto
            cantidad = item.cantidad

            stock_anterior = producto.stock_actual
            producto.stock_actual += cantidad
            producto.save()

            # Registrar movimiento de reversión
            Movimiento_inventario.objects.create(
                tipo_movimiento='ajuste',
                cantidad=cantidad,
                stock_anterior=stock_anterior,
                stock_nuevo=producto.stock_actual,
                precio_unitario=producto.precio_venta,
                valor_total=item.subtotal_linea,
                referencia_id=venta.id,
                tipo_referencia='ajuste',
                observaciones=f'Reversión por eliminación de venta {venta.numero_factura}',
                imagen_comprobante='',
                usuarios_id_usuario=request.user,
                producto_idproducto=producto,
            )

        numero_factura = venta.numero_factura
        venta.delete()

        messages.success(
            request,
            f'Venta {numero_factura} eliminada exitosamente. El stock ha sido devuelto al inventario.'
        )

    except Exception as e:
        messages.error(request, f'Error al eliminar la venta: {str(e)}')

    return redirect('ventas_listar')

# ===== VISTAS DE MENSAJERÍA (SOLO ADMIN) =====

@admin_required
def mensajeria_listar(request):
    """
    Lista todas las empresas de mensajería (solo admin)
    """
    search = request.GET.get('search', '')
    
    if search:
        mensajerias = Mensajeria.objects.filter(
            models.Q(nombre_mensajeria__icontains=search) |
            models.Q(cobertura__icontains=search)
        ).order_by('nombre_mensajeria')
    else:
        mensajerias = Mensajeria.objects.all().order_by('nombre_mensajeria')
    
    paginator = Paginator(mensajerias, 10)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    context = {
        'page_obj': page_obj,
        'search': search,
        'usuario': request.user,
        'es_admin': True
    }
    
    return render(request, 'mensajeria/listar.html', context)


@admin_required
def mensajeria_crear(request):
    """
    Crear nueva empresa de mensajería (solo admin)
    """
    if request.method == 'POST':
        form = MensajeriaForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, 'Empresa de mensajería creada exitosamente')
            return redirect('mensajeria_listar')
    else:
        form = MensajeriaForm()
    
    context = {
        'form': form,
        'titulo': 'Nueva Empresa de Mensajería',
        'usuario': request.user,
        'es_admin': True
    }
    
    return render(request, 'mensajeria/crear.html', context)


@admin_required
def mensajeria_editar(request, id):
    """
    Editar empresa de mensajería (solo admin)
    """
    mensajeria = get_object_or_404(Mensajeria, pk=id)
    
    if request.method == 'POST':
        form = MensajeriaForm(request.POST, instance=mensajeria)
        if form.is_valid():
            form.save()
            messages.success(request, 'Empresa de mensajería actualizada exitosamente')
            return redirect('mensajeria_listar')
    else:
        form = MensajeriaForm(instance=mensajeria)
    
    context = {
        'form': form,
        'titulo': 'Editar Empresa de Mensajería',
        'mensajeria': mensajeria,
        'usuario': request.user,
        'es_admin': True
    }
    
    return render(request, 'mensajeria/editar.html', context)


@admin_required
def mensajeria_eliminar(request, id):
    """
    Eliminar empresa de mensajería (solo admin)
    """
    mensajeria = get_object_or_404(Mensajeria, pk=id)
    nombre = mensajeria.nombre_mensajeria
    mensajeria.delete()
    
    messages.success(request, f'Empresa {nombre} eliminada exitosamente')
    return redirect('mensajeria_listar')


# ===== VISTAS DE ENVÍOS =====

@login_required(login_url='login')
def envios_listar(request):
    """
    Lista todos los envíos
    Admin: ve todos | Operario: ve los suyos
    + Muestra ventas sin envío asignado
    """
    search = request.GET.get('search', '')
    estado = request.GET.get('estado', '')
    search_pendientes = request.GET.get('search_pendientes', '')
    
    # Filtrar según rol
    if request.user.tipo_usu == 'administrador':
        envios = Envio.objects.all().order_by('-fecha_envio')
        ventas_query = Venta.objects.all()
    else:
        envios = Envio.objects.filter(usuarios_id_usuario=request.user).order_by('-fecha_envio')
        ventas_query = Venta.objects.filter(usuarios_id_usuario=request.user)
    
    # Aplicar filtros a envíos
    if search:
        envios = envios.filter(
            models.Q(venta_idfactura__numero_factura__icontains=search) |
            models.Q(direccion_envio__icontains=search) |
            models.Q(fk_mensajeria__nombre_mensajeria__icontains=search)
        )
    
    if estado:
        envios = envios.filter(estado_envio=estado)
    
    # Paginación de envíos
    paginator = Paginator(envios, 10)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    # ===== OBTENER VENTAS SIN ENVÍO ASIGNADO =====
    ventas_con_envio = Envio.objects.values_list('venta_idfactura', flat=True)
    ventas_sin_envio = ventas_query.exclude(id__in=ventas_con_envio).select_related('usuarios_id_usuario').order_by('-fecha_factura')
    
    # Filtrar ventas pendientes si hay búsqueda
    if search_pendientes:
        ventas_sin_envio = ventas_sin_envio.filter(numero_factura__icontains=search_pendientes)
    
    # Paginación de ventas sin envío
    paginator_pendientes = Paginator(ventas_sin_envio, 5)
    page_pendientes = request.GET.get('page_pendientes')
    ventas_pendientes_obj = paginator_pendientes.get_page(page_pendientes)
    
    context = {
        'page_obj': page_obj,
        'search': search,
        'estado': estado,
        'ventas_pendientes_obj': ventas_pendientes_obj,
        'search_pendientes': search_pendientes,
        'total_pendientes': ventas_sin_envio.count(),
        'usuario': request.user,
        'es_admin': request.user.tipo_usu == 'administrador'
    }
    
    return render(request, 'envios/listar.html', context)



@login_required(login_url='login')
def envios_crear(request):
    """
    Crear nuevo envío y notificar por correo
    """
    from django.core.mail import EmailMessage
    from django.conf import settings
    from reportlab.lib.pagesizes import letter
    from reportlab.lib import colors
    from reportlab.lib.units import inch
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.enums import TA_CENTER
    from io import BytesIO
    
    venta_id = request.GET.get('venta_id', None)

    if request.method == 'POST':
        form = EnvioForm(request.POST)
        if form.is_valid():
            envio = form.save(commit=False)
            envio.usuarios_id_usuario = request.user
            envio.save()
            
            # ===== GENERAR PDF DE LA FACTURA =====
            venta = envio.venta_idfactura
            productos_venta = Venta_has_producto.objects.filter(venta_idfactura=venta)
            
            buffer = BytesIO()
            doc = SimpleDocTemplate(buffer, pagesize=letter)
            elements = []

            styles = getSampleStyleSheet()
            title_style = ParagraphStyle(
                'CustomTitle',
                parent=styles['Heading1'],
                fontSize=24,
                textColor=colors.HexColor('#3d862e'),
                alignment=TA_CENTER
            )

            elements.append(Paragraph("ROMAR NATURAL", title_style))
            elements.append(Paragraph("NIT: 52101085", styles['Normal']))
            elements.append(Paragraph("Teléfono: 3053615676", styles['Normal']))
            elements.append(Spacer(1, 0.3 * inch))

            elements.append(Paragraph(f"<b>FACTURA: {venta.numero_factura}</b>", styles['Heading2']))
            elements.append(Paragraph(f"Fecha: {venta.fecha_factura.strftime('%d/%m/%Y %H:%M')}", styles['Normal']))
            elements.append(Paragraph(f"Vendedor: {venta.usuarios_id_usuario.nombre_completo}", styles['Normal']))

            if venta.nombre_cliente or venta.correo_cliente or venta.direccion_cliente:
                elements.append(Spacer(1, 0.2 * inch))
                elements.append(Paragraph("<b>Datos del Cliente</b>", styles['Heading3']))
                if venta.nombre_cliente:
                    elements.append(Paragraph(f"Nombre: {venta.nombre_cliente}", styles['Normal']))
                if venta.correo_cliente:
                    elements.append(Paragraph(f"Correo: {venta.correo_cliente}", styles['Normal']))
                if venta.telefono_cliente:
                    elements.append(Paragraph(f"Teléfono: {venta.telefono_cliente}", styles['Normal']))
                if venta.direccion_cliente:
                    elements.append(Paragraph(f"Dirección: {venta.direccion_cliente}", styles['Normal']))

            elements.append(Spacer(1, 0.3 * inch))

            data = [['#', 'Producto', 'Cant.', 'Precio Unit.', 'Subtotal']]
            for idx, item in enumerate(productos_venta, 1):
                data.append([
                    str(idx),
                    item.producto_idproducto.nombre_producto[:30],
                    str(item.cantidad),
                    f"${item.valor_unitario:,.0f}",
                    f"${item.subtotal_linea:,.0f}",
                ])

            table = Table(data, colWidths=[0.5 * inch, 3 * inch, 0.8 * inch, 1.2 * inch, 1.2 * inch])
            table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#3d862e')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 12),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ]))

            elements.append(table)
            elements.append(Spacer(1, 0.3 * inch))

            totales_data = [
                ['Subtotal:', f"${venta.subtotal:,.0f}"],
                ['Descuento:', f"-${venta.descuento:,.0f}"],
                ['IVA (19%):', f"${venta.iva:,.0f}"],
                ['<b>TOTAL:</b>', f"<b>${venta.valor_total:,.0f}</b>"],
            ]

            if venta.abono > 0:
                totales_data.append(['Abono:', f"${venta.abono:,.0f}"])
                totales_data.append(['<b>Saldo Pendiente:</b>', f"<b>${venta.saldo_pendiente:,.0f}</b>"])

            totales_table = Table(totales_data, colWidths=[3 * inch, 2 * inch])
            totales_table.setStyle(TableStyle([
                ('ALIGN', (0, 0), (-1, -1), 'RIGHT'),
                ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
                ('FONTSIZE', (0, -1), (-1, -1), 14),
                ('LINEABOVE', (0, -2), (-1, -2), 2, colors.black),
            ]))

            elements.append(totales_table)
            elements.append(Spacer(1, 0.5 * inch))

            elements.append(Paragraph(f"<b>Método de Pago:</b> {venta.get_metodo_pago_display()}", styles['Normal']))
            elements.append(Paragraph(f"<b>Estado:</b> {venta.get_estado_pago_display()}", styles['Normal']))

            if venta.observaciones:
                elements.append(Spacer(1, 0.2 * inch))
                elements.append(Paragraph(f"<b>Observaciones:</b> {venta.observaciones}", styles['Normal']))

            elements.append(Spacer(1, 0.5 * inch))
            elements.append(Paragraph("Gracias por su compra", styles['Normal']))

            doc.build(elements)
            pdf_content = buffer.getvalue()
            buffer.close()

            # Mapeo de estados a texto amigable
            estados_texto = {
                'pendiente': 'Pendiente de envío',
                'en_transito': 'En tránsito',
                'entregado': 'Entregado',
                'devuelto': 'Devuelto'
            }
            estado_display = estados_texto.get(envio.estado_envio, envio.estado_envio)

            # ===== ENVIAR CORREO AL CLIENTE =====
            if venta.correo_cliente:
                try:
                    email_cliente = EmailMessage(
                        subject=f'Estado de tu envío - {venta.numero_factura}',
                        body=(
                            f'Hola {venta.nombre_cliente},\n\n'
                            f'Tu pedido ha sido registrado para envío.\n\n'
                            f'Estado actual: {estado_display}\n'
                            f'Factura: {venta.numero_factura}\n'
                            f'Fecha de envío: {envio.fecha_envio.strftime("%d/%m/%Y")}\n'
                            f'Fecha estimada de entrega: {envio.fecha_entrega.strftime("%d/%m/%Y")}\n'
                            f'Dirección de entrega: {envio.direccion_envio}\n'
                            f'Empresa de mensajería: {envio.fk_mensajeria.nombre_mensajeria}\n\n'
                            f'Adjuntamos tu factura en formato PDF.\n\n'
                            f'Gracias por tu preferencia.\n\n'
                            f'Romar Natural'
                        ),
                        from_email=settings.DEFAULT_FROM_EMAIL,
                        to=[venta.correo_cliente],
                    )
                    email_cliente.attach(f'Factura_{venta.numero_factura}.pdf', pdf_content, 'application/pdf')
                    email_cliente.send(fail_silently=True)
                except Exception:
                    pass

            # ===== ENVIAR CORREO AL RESPONSABLE DEL ENVÍO =====
            if envio.usuarios_id_usuario.correo:
                try:
                    email_responsable = EmailMessage(
                        subject=f'Envío asignado - {venta.numero_factura}',
                        body=(
                            f'Hola {envio.usuarios_id_usuario.nombre_completo},\n\n'
                            f'Se ha registrado un nuevo envío bajo tu responsabilidad.\n\n'
                            f'Factura: {venta.numero_factura}\n'
                            f'Cliente: {venta.nombre_cliente or "N/A"}\n'
                            f'Estado actual: {estado_display}\n'
                            f'Fecha de envío: {envio.fecha_envio.strftime("%d/%m/%Y")}\n'
                            f'Fecha estimada de entrega: {envio.fecha_entrega.strftime("%d/%m/%Y")}\n'
                            f'Dirección: {envio.direccion_envio}\n\n'
                            f'RECORDATORIO: Por favor actualiza el estado del envío según avance el proceso.\n'
                            f'Accede al sistema para registrar novedades o cambios de estado.\n\n'
                            f'Adjuntamos la factura en formato PDF.\n\n'
                            f'Romar Natural'
                        ),
                        from_email=settings.DEFAULT_FROM_EMAIL,
                        to=[envio.usuarios_id_usuario.correo],
                    )
                    email_responsable.attach(f'Factura_{venta.numero_factura}.pdf', pdf_content, 'application/pdf')
                    email_responsable.send(fail_silently=True)
                except Exception:
                    pass

            messages.success(
                request,
                f'Envío registrado exitosamente para la venta {envio.venta_idfactura.numero_factura}'
            )
            return redirect('envios_detalle', id=envio.id)
    else:
        if venta_id:
            try:
                venta = Venta.objects.get(id=venta_id)

                if Envio.objects.filter(venta_idfactura=venta).exists():
                    messages.warning(
                        request,
                        f'La venta {venta.numero_factura} ya tiene un envío asociado.'
                    )
                    return redirect('ventas_listar')

                initial_data = {
                    'venta_idfactura': venta,
                    'direccion_envio': venta.direccion_cliente or '',
                }
                form = EnvioForm(initial=initial_data)
            except Venta.DoesNotExist:
                messages.error(request, 'Venta no encontrada')
                form = EnvioForm()
        else:
            form = EnvioForm()

    context = {
        'form': form,
        'titulo': 'Registrar Nuevo Envío',
        'usuario': request.user,
        'es_admin': request.user.tipo_usu == 'administrador',
    }

    return render(request, 'envios/crear.html', context)

@login_required(login_url='login')
def envios_editar(request, id):
    """
    Editar envío existente y notificar cambios
    Admin: puede editar todo
    Operario: solo puede editar estado y novedades
    """
    from django.core.mail import EmailMessage
    from django.conf import settings
    from reportlab.lib.pagesizes import letter
    from reportlab.lib import colors
    from reportlab.lib.units import inch
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.enums import TA_CENTER
    from io import BytesIO
    
    envio = get_object_or_404(Envio, pk=id)
    es_admin = request.user.tipo_usu == 'administrador'

    if request.method == 'POST':
        # Guardar estado anterior para detectar cambios
        estado_anterior = envio.estado_envio
        
        if es_admin:
            form = EnvioForm(request.POST, instance=envio)
        else:
            form = EnvioEditarOperarioForm(request.POST, instance=envio)

        if form.is_valid():
            envio = form.save()
            
            # Detectar si cambió el estado
            estado_cambio = estado_anterior != envio.estado_envio
            
            if estado_cambio:
                # ===== GENERAR PDF DE LA FACTURA =====
                venta = envio.venta_idfactura
                productos_venta = Venta_has_producto.objects.filter(venta_idfactura=venta)
                
                buffer = BytesIO()
                doc = SimpleDocTemplate(buffer, pagesize=letter)
                elements = []

                styles = getSampleStyleSheet()
                title_style = ParagraphStyle(
                    'CustomTitle',
                    parent=styles['Heading1'],
                    fontSize=24,
                    textColor=colors.HexColor('#3d862e'),
                    alignment=TA_CENTER
                )

                elements.append(Paragraph("ROMAR NATURAL", title_style))
                elements.append(Paragraph("NIT: 52101085", styles['Normal']))
                elements.append(Paragraph("Teléfono: 3053615676", styles['Normal']))
                elements.append(Spacer(1, 0.3 * inch))

                elements.append(Paragraph(f"<b>FACTURA: {venta.numero_factura}</b>", styles['Heading2']))
                elements.append(Paragraph(f"Fecha: {venta.fecha_factura.strftime('%d/%m/%Y %H:%M')}", styles['Normal']))
                elements.append(Paragraph(f"Vendedor: {venta.usuarios_id_usuario.nombre_completo}", styles['Normal']))

                if venta.nombre_cliente or venta.correo_cliente or venta.direccion_cliente:
                    elements.append(Spacer(1, 0.2 * inch))
                    elements.append(Paragraph("<b>Datos del Cliente</b>", styles['Heading3']))
                    if venta.nombre_cliente:
                        elements.append(Paragraph(f"Nombre: {venta.nombre_cliente}", styles['Normal']))
                    if venta.correo_cliente:
                        elements.append(Paragraph(f"Correo: {venta.correo_cliente}", styles['Normal']))
                    if venta.telefono_cliente:
                        elements.append(Paragraph(f"Teléfono: {venta.telefono_cliente}", styles['Normal']))
                    if venta.direccion_cliente:
                        elements.append(Paragraph(f"Dirección: {venta.direccion_cliente}", styles['Normal']))

                elements.append(Spacer(1, 0.3 * inch))

                data = [['#', 'Producto', 'Cant.', 'Precio Unit.', 'Subtotal']]
                for idx, item in enumerate(productos_venta, 1):
                    data.append([
                        str(idx),
                        item.producto_idproducto.nombre_producto[:30],
                        str(item.cantidad),
                        f"${item.valor_unitario:,.0f}",
                        f"${item.subtotal_linea:,.0f}",
                    ])

                table = Table(data, colWidths=[0.5 * inch, 3 * inch, 0.8 * inch, 1.2 * inch, 1.2 * inch])
                table.setStyle(TableStyle([
                    ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#3d862e')),
                    ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                    ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                    ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                    ('FONTSIZE', (0, 0), (-1, 0), 12),
                    ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                    ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                    ('GRID', (0, 0), (-1, -1), 1, colors.black),
                ]))

                elements.append(table)
                elements.append(Spacer(1, 0.3 * inch))

                totales_data = [
                    ['Subtotal:', f"${venta.subtotal:,.0f}"],
                    ['Descuento:', f"-${venta.descuento:,.0f}"],
                    ['IVA (19%):', f"${venta.iva:,.0f}"],
                    ['<b>TOTAL:</b>', f"<b>${venta.valor_total:,.0f}</b>"],
                ]

                if venta.abono > 0:
                    totales_data.append(['Abono:', f"${venta.abono:,.0f}"])
                    totales_data.append(['<b>Saldo Pendiente:</b>', f"<b>${venta.saldo_pendiente:,.0f}</b>"])

                totales_table = Table(totales_data, colWidths=[3 * inch, 2 * inch])
                totales_table.setStyle(TableStyle([
                    ('ALIGN', (0, 0), (-1, -1), 'RIGHT'),
                    ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
                    ('FONTSIZE', (0, -1), (-1, -1), 14),
                    ('LINEABOVE', (0, -2), (-1, -2), 2, colors.black),
                ]))

                elements.append(totales_table)
                elements.append(Spacer(1, 0.5 * inch))

                elements.append(Paragraph(f"<b>Método de Pago:</b> {venta.get_metodo_pago_display()}", styles['Normal']))
                elements.append(Paragraph(f"<b>Estado:</b> {venta.get_estado_pago_display()}", styles['Normal']))

                if venta.observaciones:
                    elements.append(Spacer(1, 0.2 * inch))
                    elements.append(Paragraph(f"<b>Observaciones:</b> {venta.observaciones}", styles['Normal']))

                elements.append(Spacer(1, 0.5 * inch))
                elements.append(Paragraph("Gracias por su compra", styles['Normal']))

                doc.build(elements)
                pdf_content = buffer.getvalue()
                buffer.close()

                # Mapeo de estados
                estados_texto = {
                    'pendiente': 'Pendiente de envío',
                    'en_transito': 'En tránsito',
                    'entregado': 'Entregado',
                    'devuelto': 'Devuelto'
                }
                estado_display = estados_texto.get(envio.estado_envio, envio.estado_envio)

                # ===== ENVIAR CORREO AL CLIENTE =====
                if venta.correo_cliente:
                    try:
                        email_cliente = EmailMessage(
                            subject=f'Actualización de tu envío - {venta.numero_factura}',
                            body=(
                                f'Hola {venta.nombre_cliente},\n\n'
                                f'El estado de tu envío ha sido actualizado.\n\n'
                                f'Estado actual: {estado_display}\n'
                                f'Factura: {venta.numero_factura}\n'
                                f'Fecha estimada de entrega: {envio.fecha_entrega.strftime("%d/%m/%Y")}\n'
                                f'Dirección de entrega: {envio.direccion_envio}\n\n'
                                f'{f"Novedades: {envio.novedades}" if envio.novedades else ""}\n\n'
                                f'Adjuntamos tu factura en formato PDF.\n\n'
                                f'Gracias por tu preferencia.\n\n'
                                f'Romar Natural'
                            ),
                            from_email=settings.DEFAULT_FROM_EMAIL,
                            to=[venta.correo_cliente],
                        )
                        email_cliente.attach(f'Factura_{venta.numero_factura}.pdf', pdf_content, 'application/pdf')
                        email_cliente.send(fail_silently=True)
                    except Exception:
                        pass

                # ===== ENVIAR CORREO AL RESPONSABLE DEL ENVÍO =====
                if envio.usuarios_id_usuario.correo:
                    try:
                        email_responsable = EmailMessage(
                            subject=f'Estado actualizado - {venta.numero_factura}',
                            body=(
                                f'Hola {envio.usuarios_id_usuario.nombre_completo},\n\n'
                                f'El estado del envío ha sido actualizado.\n\n'
                                f'Factura: {venta.numero_factura}\n'
                                f'Cliente: {venta.nombre_cliente or "N/A"}\n'
                                f'Estado actual: {estado_display}\n'
                                f'Fecha estimada de entrega: {envio.fecha_entrega.strftime("%d/%m/%Y")}\n\n'
                                f'RECORDATORIO: Mantén actualizado el estado y registra cualquier novedad.\n\n'
                                f'Adjuntamos la factura en formato PDF.\n\n'
                                f'Romar Natural'
                            ),
                            from_email=settings.DEFAULT_FROM_EMAIL,
                            to=[envio.usuarios_id_usuario.correo],
                        )
                        email_responsable.attach(f'Factura_{venta.numero_factura}.pdf', pdf_content, 'application/pdf')
                        email_responsable.send(fail_silently=True)
                    except Exception:
                        pass

            messages.success(request, 'Envío actualizado exitosamente')
            return redirect('envios_detalle', id=envio.id)
    else:
        if es_admin:
            initial_data = {}
            if not envio.direccion_envio and envio.venta_idfactura and envio.venta_idfactura.direccion_cliente:
                initial_data['direccion_envio'] = envio.venta_idfactura.direccion_cliente

            form = EnvioForm(instance=envio, initial=initial_data)
        else:
            form = EnvioEditarOperarioForm(instance=envio)

    context = {
        'form': form,
        'titulo': 'Editar Envío',
        'envio': envio,
        'usuario': request.user,
        'es_admin': es_admin,
        'solo_estado': not es_admin,
    }

    return render(request, 'envios/editar.html', context)

@admin_required
def envios_eliminar(request, id):
    """
    Eliminar envío (solo admin)
    """
    envio = get_object_or_404(Envio, pk=id)
    factura = envio.venta_idfactura.numero_factura
    envio.delete()
    
    messages.success(request, f'Envío de la factura {factura} eliminado exitosamente')
    return redirect('envios_listar')


@login_required(login_url='login')
def envios_detalle(request, id):
    """
    Ver detalles de un envío
    """
    envio = get_object_or_404(Envio, pk=id)
    
    context = {
        'envio': envio,
        'usuario': request.user,
        'es_admin': request.user.tipo_usu == 'administrador'
    }
    
    return render(request, 'envios/detalle.html', context)
